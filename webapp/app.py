#!/usr/bin/env python3
"""Leadway Health Analytics Web App — Upload CSV, get instant report."""

import os
import sys
import uuid
import base64
import json
from pathlib import Path
from functools import wraps
from werkzeug.security import generate_password_hash, check_password_hash
from flask import Flask, render_template, request, redirect, url_for, send_file, session, flash

sys.path.insert(0, str(Path(__file__).resolve().parent.parent))

app = Flask(__name__)
app.config["UPLOAD_FOLDER"] = Path(__file__).parent / "uploads"
app.config["MAX_CONTENT_LENGTH"] = 100 * 1024 * 1024  # 100MB
app.secret_key = os.environ.get("SECRET_KEY", "leadway-health-analytics-2026")

BASE_DIR = Path(__file__).resolve().parent.parent
REPORTS_DIR = BASE_DIR / "reports"
LOGO_PATH = BASE_DIR / "leadway health logo 20266.jpg"
USERS_FILE = Path(__file__).parent / "users.json"


def load_users():
    if USERS_FILE.exists():
        return json.loads(USERS_FILE.read_text())
    # Default admin user — change password on first login
    default = {
        "admin@leadwayhealth.com": {
            "password": generate_password_hash("admin123"),
            "name": "Admin",
            "role": "admin"
        }
    }
    USERS_FILE.write_text(json.dumps(default, indent=2))
    return default


def save_users(users):
    USERS_FILE.write_text(json.dumps(users, indent=2))


def login_required(f):
    @wraps(f)
    def decorated(*args, **kwargs):
        if "user" not in session:
            return redirect(url_for("login"))
        return f(*args, **kwargs)
    return decorated


def admin_required(f):
    @wraps(f)
    def decorated(*args, **kwargs):
        if "user" not in session:
            return redirect(url_for("login"))
        users = load_users()
        user = users.get(session["user"], {})
        if user.get("role") != "admin":
            flash("Admin access required.")
            return redirect(url_for("index"))
        return f(*args, **kwargs)
    return decorated


def get_logo_b64():
    if LOGO_PATH.exists():
        return base64.b64encode(LOGO_PATH.read_bytes()).decode()
    return ""


@app.route("/login", methods=["GET", "POST"])
def login():
    logo = get_logo_b64()
    if request.method == "POST":
        email = request.form.get("email", "").strip().lower()
        password = request.form.get("password", "")
        users = load_users()
        user = users.get(email)
        if user and check_password_hash(user["password"], password):
            session["user"] = email
            session["name"] = user.get("name", email.split("@")[0])
            return redirect(url_for("index"))
        flash("Invalid email or password.")
    return render_template("login.html", logo=logo)


@app.route("/logout")
def logout():
    session.clear()
    return redirect(url_for("login"))


@app.route("/manage-users", methods=["GET", "POST"])
@admin_required
def manage_users():
    logo = get_logo_b64()
    users = load_users()
    if request.method == "POST":
        action = request.form.get("action")
        if action == "add":
            email = request.form.get("email", "").strip().lower()
            name = request.form.get("name", "").strip()
            password = request.form.get("password", "")
            role = request.form.get("role", "user")
            if email and password:
                users[email] = {
                    "password": generate_password_hash(password),
                    "name": name or email.split("@")[0],
                    "role": role
                }
                save_users(users)
                flash(f"User {email} added.")
        elif action == "remove":
            email = request.form.get("email", "").strip().lower()
            if email in users and email != session["user"]:
                del users[email]
                save_users(users)
                flash(f"User {email} removed.")
    user_list = [{"email": e, "name": u["name"], "role": u["role"]} for e, u in users.items()]
    return render_template("manage_users.html", logo=logo, users=user_list)


def list_reports(prefix="", ext="html"):
    """List reports, optionally filtered by filename prefix and extension."""
    if not REPORTS_DIR.exists():
        return []
    all_reports = sorted(REPORTS_DIR.glob(f"*.{ext}"), key=lambda f: f.stat().st_mtime, reverse=True)
    if prefix:
        all_reports = [r for r in all_reports if r.name.lower().startswith(prefix.lower())]
    return [{"name": r.stem.replace("_", " "), "file": r.name} for r in all_reports]


@app.route("/")
@login_required
def index():
    logo = get_logo_b64()
    user_name = session.get("name", "")
    users = load_users()
    is_admin = users.get(session.get("user", ""), {}).get("role") == "admin"
    report_list = list_reports()
    return render_template("dashboard.html", logo=logo, reports=report_list, user_name=user_name, is_admin=is_admin)


# ═══════════════════════════════════════════════
# MODULE 1 — Utilization Report
# ═══════════════════════════════════════════════

@app.route("/utilization")
@login_required
def utilization():
    logo = get_logo_b64()
    report_list = list_reports()
    return render_template("utilization.html", logo=logo, reports=report_list)


@app.route("/utilization/upload", methods=["POST"])
@login_required
def utilization_upload():
    # Get form data
    brokered = request.form.get("brokered", "no")
    broker_fee = float(request.form.get("broker_fee", 0)) if brokered == "yes" else 0
    additional_reqs = request.form.get("additional_reqs", "").strip()

    if brokered == "yes":
        admin_pct = 12.0
    else:
        admin_pct = 15.0
    nhia_pct = 2.0

    # Handle file uploads
    files = request.files.getlist("files")
    if not files or all(f.filename == "" for f in files):
        return redirect(url_for("utilization"))

    # Save uploaded files
    session_id = str(uuid.uuid4())[:8]
    upload_dir = app.config["UPLOAD_FOLDER"] / session_id
    upload_dir.mkdir(parents=True, exist_ok=True)

    saved_files = {}
    for f in files:
        if f.filename:
            fname = f.filename.lower()
            fpath = upload_dir / f.filename
            f.save(fpath)
            if "claim" in fname:
                saved_files["claims"] = fpath
            elif "production" in fname or "premium" in fname:
                saved_files["production"] = fpath
            elif "benefit" in fname:
                saved_files["benefit"] = fpath
            elif "hospital" in fname:
                saved_files["hospital"] = fpath
            else:
                saved_files[f.filename] = fpath

    if "claims" not in saved_files:
        return "No claims file found. Please upload a file with 'claims' in the name.", 400

    # Generate report
    try:
        report_path = generate_report_from_files(
            saved_files, admin_pct, nhia_pct, broker_fee, additional_reqs, session_id
        )
        return redirect(url_for("view_report", filename=report_path.name))
    except Exception as e:
        import traceback
        return f"<pre>Error generating report:\n{traceback.format_exc()}</pre>", 500


def generate_report_from_files(files, admin_pct, nhia_pct, broker_fee, additional_reqs, session_id):
    """Generate an HTML analytics report from uploaded files."""
    import pandas as pd
    import numpy as np

    REPORTS_DIR.mkdir(parents=True, exist_ok=True)

    # Load claims
    claims_path = files["claims"]
    if str(claims_path).endswith(".csv"):
        claims = pd.read_csv(claims_path)
    else:
        claims = pd.read_excel(claims_path)

    claims.columns = claims.columns.str.strip()

    # Standardize column names
    col_map = {}
    for col in claims.columns:
        cl = col.lower().strip()
        if "claim" in cl and ("num" in cl or "no" in cl):
            col_map[col] = "Claim_Number"
        elif cl in ("amt paid", "amount paid", "amountpaid"):
            col_map[col] = "Amount_Paid"
        elif cl in ("amt claimed", "amount claimed", "amountclaimed"):
            col_map[col] = "Amount_Claimed"
        elif "treatment" in cl and "date" in cl:
            col_map[col] = "Treatment_Date"
        elif cl in ("received on", "receivedon", "received_on"):
            col_map[col] = "Received_Date"
        elif "claim" in cl and "status" in cl:
            col_map[col] = "Claim_Status"
        elif cl in ("member ship no", "membershipno", "member_id", "enrollee id", "member enrollee id"):
            col_map[col] = "Member_ID"
        elif cl in ("department",):
            col_map[col] = "Department"
        elif cl in ("provider",):
            col_map[col] = "Provider"
        elif cl in ("principal member", "principalmember"):
            col_map[col] = "Principal_Member"
        elif cl in ("description",):
            col_map[col] = "Description"
        elif cl in ("scheme",):
            col_map[col] = "Scheme"
        elif cl in ("groupcode", "group code", "group_code"):
            col_map[col] = "Group_Code"
        elif cl in ("currentage", "age"):
            col_map[col] = "Age"
        elif "diagnosis" in cl and "desc" in cl:
            col_map[col] = "Diagnosis_Description"

    claims.rename(columns=col_map, inplace=True)

    # Parse numerics
    for col in ["Amount_Paid", "Amount_Claimed", "Age"]:
        if col in claims.columns:
            claims[col] = pd.to_numeric(claims[col].astype(str).str.replace(",", ""), errors="coerce")

    # Parse dates — handle string dates, Excel serial numbers, and mixed formats
    for col in ["Treatment_Date", "Received_Date"]:
        if col in claims.columns:
            raw_series = claims[col].copy()
            # First check if values are numeric (Excel serial numbers)
            numeric_vals = pd.to_numeric(raw_series, errors="coerce")
            numeric_count = numeric_vals.notna().sum()
            excel_like = numeric_vals.notna() & (numeric_vals > 30000) & (numeric_vals < 60000)
            if excel_like.sum() > numeric_count * 0.5 and numeric_count > 0:
                # Excel serial number dates (days since 1899-12-30)
                claims[col] = pd.NaT
                claims.loc[excel_like, col] = pd.to_datetime("1899-12-30") + pd.to_timedelta(numeric_vals[excel_like], unit="D")
            else:
                # Try standard datetime parsing with dayfirst
                claims[col] = pd.to_datetime(raw_series, errors="coerce", dayfirst=True)
                # If that failed for most rows, try other formats
                if claims[col].notna().sum() < len(claims) * 0.3:
                    claims[col] = pd.to_datetime(raw_series, errors="coerce", format="mixed", dayfirst=True)

    # Load production if available
    prod = None
    if "production" in files:
        prod_path = files["production"]
        if str(prod_path).endswith(".csv"):
            prod = pd.read_csv(prod_path)
        else:
            prod = pd.read_excel(prod_path)
        prod.columns = prod.columns.str.strip()

        # Standardize prod columns
        prod_col_map = {}
        for col in prod.columns:
            cl = col.lower().strip()
            if cl in ("l", "member enrollee id", "member_enrolleeid", "memberenrolleeid"):
                prod_col_map[col] = "Member_ID"
            elif "premium" in cl and ("fee" in cl or "individual" in cl):
                prod_col_map[col] = "Premium"
            elif "effective" in cl:
                prod_col_map[col] = "Effective_Date"
            elif "expiry" in cl or "expir" in cl:
                prod_col_map[col] = "Expiry_Date"
            elif "status" in cl and "desc" in cl:
                prod_col_map[col] = "Status"
            elif "relationship" in cl:
                prod_col_map[col] = "Relationship"
            elif "gender" in cl:
                prod_col_map[col] = "Gender"
            elif "plan" in cl:
                prod_col_map[col] = "Plan"
            elif "customer" in cl or "client" in cl and "name" in cl:
                prod_col_map[col] = "Client_Name"
        prod.rename(columns=prod_col_map, inplace=True)

        if "Premium" in prod.columns:
            prod["Premium"] = pd.to_numeric(prod["Premium"].astype(str).str.replace(",", ""), errors="coerce")
        for col in ["Effective_Date", "Expiry_Date"]:
            if col in prod.columns:
                prod[col] = pd.to_datetime(prod[col], errors="coerce")

    # Detect client name
    client_name = "Client"
    if "Group_Code" in claims.columns:
        client_name = claims["Group_Code"].mode().iloc[0] if not claims["Group_Code"].mode().empty else "Client"
    elif prod is not None and "Client_Name" in prod.columns:
        client_name = prod["Client_Name"].mode().iloc[0] if not prod["Client_Name"].mode().empty else "Client"

    # ── Metrics ──
    unique_claims = int(claims["Claim_Number"].nunique()) if "Claim_Number" in claims.columns else len(claims)
    unique_members = int(claims["Member_ID"].nunique()) if "Member_ID" in claims.columns else 0
    total_enrolled = len(prod) if prod is not None else unique_members

    paid = claims[claims["Claim_Status"] == "Paid Claims"] if "Claim_Status" in claims.columns else claims
    pipeline = claims[claims["Claim_Status"].isin(["Awaiting Payment", "Claims for adjudication", "In Process"])] if "Claim_Status" in claims.columns else pd.DataFrame()
    paid_total = float(paid["Amount_Paid"].sum()) if "Amount_Paid" in paid.columns else 0
    pipeline_total = float(pipeline["Amount_Claimed"].sum()) if "Amount_Claimed" in pipeline.columns and not pipeline.empty else 0

    # Earned premium
    earned_total = 0
    written_total = 0
    if prod is not None and "Premium" in prod.columns and "Effective_Date" in prod.columns and "Expiry_Date" in prod.columns:
        as_of = pd.Timestamp.now().normalize()
        ep = prod.dropna(subset=["Effective_Date", "Expiry_Date", "Premium"]).copy()
        ep = ep[ep["Premium"] > 0]
        total_days = (ep["Expiry_Date"] - ep["Effective_Date"]).dt.days
        elapsed = (np.minimum(as_of, ep["Expiry_Date"]) - ep["Effective_Date"]).dt.days.clip(lower=0)
        frac = (elapsed / total_days).clip(upper=1.0)
        ep["Earned"] = ep["Premium"] * frac
        earned_total = float(ep["Earned"].sum())
        written_total = float(ep["Premium"].sum())

    # Simple IBNR
    ibnr_total = 0
    if "Treatment_Date" in claims.columns and "Received_Date" in claims.columns:
        df_i = claims.dropna(subset=["Treatment_Date", "Received_Date"]).copy()
        monthly_paid = df_i.groupby(df_i["Treatment_Date"].dt.to_period("M"))["Amount_Paid"].sum()
        if len(monthly_paid) > 2:
            avg_monthly = monthly_paid.iloc[:-1].median()
            ibnr_total = float(max(avg_monthly * 0.3, 0))

    total_incurred = float(paid_total + pipeline_total + ibnr_total)
    mlr_pct = (total_incurred / earned_total * 100) if earned_total > 0 else 0
    cor_pct = mlr_pct + admin_pct + nhia_pct
    admin_amount = earned_total * admin_pct / 100
    nhia_amount = earned_total * nhia_pct / 100

    avg_per_member = paid_total / unique_members if unique_members > 0 else 0
    avg_per_visit = paid_total / unique_claims if unique_claims > 0 else 0

    # ── Helper functions ──
    def to_initials(name):
        if not name or str(name).strip() == "":
            return ""
        parts = str(name).strip().split()
        return " ".join(p[0] + "." for p in parts if p)

    def fmt(x):
        if abs(x) >= 1_000_000:
            return f"&#8358;{x/1_000_000:,.2f}M"
        if abs(x) >= 1_000:
            return f"&#8358;{x/1_000:,.0f}K"
        return f"&#8358;{x:,.0f}"

    def fmt_full(x):
        return f"&#8358;{x:,.2f}"

    def pct(x):
        return f"{x:.1f}%"

    def month_label(period):
        return period.to_timestamp().strftime("%B %Y")

    mlr_cls = "danger" if mlr_pct > 100 else ("warning" if mlr_pct > 85 else "good")

    # ── Monthly Trend ──
    monthly_rows = ""
    if "Treatment_Date" in claims.columns:
        claims["Month"] = claims["Treatment_Date"].dt.to_period("M")
        monthly = claims.dropna(subset=["Treatment_Date"]).groupby("Month").agg(
            Claims=("Claim_Number", "nunique") if "Claim_Number" in claims.columns else ("Amount_Paid", "count"),
            Paid=("Amount_Paid", "sum"),
            Members=("Member_ID", "nunique") if "Member_ID" in claims.columns else ("Amount_Paid", "count"),
        ).sort_index()
        for period, row in monthly.iterrows():
            m_claims = row["Claims"]
            m_members = row["Members"]
            m_paid = row["Paid"]
            m_avg = m_paid / m_members if m_members > 0 else 0
            m_visits = m_claims / m_members if m_members > 0 else 0
            monthly_rows += f'<tr><td>{month_label(period)}</td><td class="num">{m_claims:,}</td><td class="num">{m_members:,}</td><td class="num">{fmt_full(m_paid)}</td><td class="num">{fmt_full(m_avg)}</td><td class="num">{m_visits:.1f}</td></tr>'

    # ── Top Providers ──
    prov_rows = ""
    if "Provider" in claims.columns:
        top_prov = claims.groupby("Provider").agg(Paid=("Amount_Paid", "sum"), Claims=("Claim_Number", "nunique") if "Claim_Number" in claims.columns else ("Amount_Paid", "count")).sort_values("Paid", ascending=False).head(10)
        top_prov["Pct"] = (top_prov["Paid"] / paid_total * 100).round(1)
        for i, (name, row) in enumerate(top_prov.iterrows(), 1):
            prov_rows += f'<tr><td>{i}</td><td>{str(name).strip().title()}</td><td class="num">{fmt_full(row["Paid"])}</td><td class="num">{pct(row["Pct"])}</td><td class="num">{row["Claims"]:,}</td></tr>'

    # ── Department Breakdown ──
    dept_rows = ""
    if "Department" in claims.columns:
        dept = claims.groupby("Department").agg(Paid=("Amount_Paid", "sum"), Claims=("Claim_Number", "nunique") if "Claim_Number" in claims.columns else ("Amount_Paid", "count")).sort_values("Paid", ascending=False)
        dept["Pct"] = (dept["Paid"] / paid_total * 100).round(1)
        for i, (name, row) in enumerate(dept.head(15).iterrows(), 1):
            dept_rows += f'<tr><td>{i}</td><td>{name}</td><td class="num">{fmt_full(row["Paid"])}</td><td class="num">{pct(row["Pct"])}</td><td class="num">{row["Claims"]:,}</td></tr>'

    # ── Disease Category (Diagnosis Bands) ──
    disease_cat_rows = ""
    if "Diagnosis_Description" in claims.columns:
        disease_categories = {
            "Malaria": ["malaria", "plasmodium"],
            "Respiratory Infections": ["respiratory", "pneumonia", "bronchitis", "asthma", "cough", "influenza", "flu", "pharyngitis", "sinusitis", "tonsillitis", "rhinitis", "laryngitis", "upper respiratory", "lower respiratory", "urt", "lrt", "urti", "lrti", "nasal"],
            "Cardiovascular & Hypertension": ["hypertension", "hypertensive", "cardiac", "heart", "cardiovascular", "angina", "stroke", "cerebrovascular", "ischaemic", "ischemic", "coronary", "arrhythmia", "blood pressure"],
            "Diabetes & Metabolic": ["diabet", "glucose", "metabolic", "thyroid", "cholesterol", "lipid", "obesity", "gout", "hyperglycaemia", "hypoglycaemia"],
            "Gastrointestinal": ["gastro", "gastritis", "ulcer", "diarrh", "dyspepsia", "colitis", "intestin", "abdominal", "stomach", "bowel", "hepatitis", "liver", "gerd", "reflux", "constipation", "appendic", "hernia", "pancreat", "gallstone", "peptic"],
            "Musculoskeletal": ["arthritis", "osteo", "muscul", "joint", "back pain", "spine", "spinal", "fracture", "rheumat", "ortho", "lumbar", "cervical", "spondyl", "musculo"],
            "Eye & Ophthalmology": ["eye", "ophthalm", "visual", "cataract", "glaucoma", "conjunctiv", "retina", "myopia", "optical", "vision"],
            "Dental & Oral": ["dental", "tooth", "teeth", "oral", "gingiv", "caries", "periodon", "dentition"],
            "Obstetric & Maternity": ["pregnan", "matern", "obstetric", "antenatal", "postnatal", "delivery", "caesarean", "c-section", "labour", "labor", "gestation", "neonatal", "perinatal"],
            "Surgery & Procedures": ["surgery", "surgical", "operation", "procedure", "excision", "repair", "implant", "biopsy", "laparoscop", "endoscop"],
            "Dermatology & Skin": ["skin", "dermat", "eczema", "rash", "wound", "abscess", "cellulitis", "fungal", "urticaria", "psoriasis", "acne"],
            "Genitourinary": ["urinary", "kidney", "renal", "bladder", "urin", "prostat", "uti", "nephri", "genital", "pelvic"],
            "Infections & Parasitic": ["typhoid", "infection", "sepsis", "hiv", "tuberculosis", "measles", "chicken pox", "viral", "bacterial", "fever", "parasit"],
            "Oncology": ["cancer", "tumour", "tumor", "malignan", "carcinoma", "oncolog", "leukaemia", "leukemia", "lymphoma", "neoplasm"],
            "Mental Health & Neurology": ["mental", "depression", "anxiety", "psychi", "epilep", "seizure", "migraine", "headache", "neuro", "insomnia", "bipolar"],
            "ENT": ["ear", "nose", "throat", "otitis", "hearing", "vertigo", "adenoid"],
        }

        def classify_diagnosis(desc):
            if not desc or str(desc).strip() == "":
                return "Other / Unclassified"
            desc_lower = str(desc).lower()
            for category, keywords in disease_categories.items():
                for kw in keywords:
                    if kw in desc_lower:
                        return category
            return "Other / Unclassified"

        claims["Disease_Category"] = claims["Diagnosis_Description"].apply(classify_diagnosis)
        disease_agg = claims.groupby("Disease_Category").agg(
            Paid=("Amount_Paid", "sum"),
            Claims=("Claim_Number", "nunique") if "Claim_Number" in claims.columns else ("Amount_Paid", "count"),
        ).sort_values("Paid", ascending=False)
        disease_agg["Pct"] = (disease_agg["Paid"] / paid_total * 100).round(1)
        for i, (cat, row) in enumerate(disease_agg.head(10).iterrows(), 1):
            disease_cat_rows += f'<tr><td>{i}</td><td>{cat}</td><td class="num">{fmt_full(row["Paid"])}</td><td class="num">{pct(row["Pct"])}</td><td class="num">{row["Claims"]:,}</td></tr>'

    # ── MLR by Age Group ──
    age_mlr_rows = ""
    if "Age" in claims.columns and earned_total > 0:
        age_bins = [0, 5, 18, 30, 40, 50, 60, 70, 120]
        age_labels = ["0-5", "6-18", "19-30", "31-40", "41-50", "51-60", "61-70", "71+"]
        claims["Age_Group"] = pd.cut(claims["Age"].dropna(), bins=age_bins, labels=age_labels, right=True)
        age_agg = claims.dropna(subset=["Age_Group"]).groupby("Age_Group", observed=True).agg(
            Paid=("Amount_Paid", "sum"),
            Claims=("Claim_Number", "nunique") if "Claim_Number" in claims.columns else ("Amount_Paid", "count"),
            Members=("Member_ID", "nunique") if "Member_ID" in claims.columns else ("Amount_Paid", "count"),
        )
        # Distribute earned premium proportionally by member count for age-group MLR
        total_age_members = age_agg["Members"].sum()
        for grp, row in age_agg.iterrows():
            if row["Members"] == 0:
                continue
            grp_earned = earned_total * (row["Members"] / total_age_members) if total_age_members > 0 else 0
            grp_mlr = (row["Paid"] / grp_earned * 100) if grp_earned > 0 else 0
            grp_avg = row["Paid"] / row["Members"] if row["Members"] > 0 else 0
            mlr_class = "danger" if grp_mlr > 100 else ("warning" if grp_mlr > 85 else "good")
            age_mlr_rows += f'<tr><td>{grp}</td><td class="num">{row["Members"]:,}</td><td class="num">{row["Claims"]:,}</td><td class="num">{fmt_full(row["Paid"])}</td><td class="num">{fmt_full(grp_earned)}</td><td class="num {mlr_class}">{pct(grp_mlr)}</td><td class="num">{fmt_full(grp_avg)}</td></tr>'

    # ── Early High Claimers ──
    early_claimers_rows = ""
    early_claimers_count = 0
    if prod is not None and "Effective_Date" in prod.columns and "Member_ID" in prod.columns and "Member_ID" in claims.columns:
        # Build member join dates and premium from production data
        member_join = prod.dropna(subset=["Effective_Date", "Member_ID"]).groupby("Member_ID").agg(
            Join_Date=("Effective_Date", "min"),
            Premium=("Premium", "sum") if "Premium" in prod.columns else ("Effective_Date", "count"),
        )
        # Get first claim date and total spent per member
        member_claims = claims.dropna(subset=["Treatment_Date"]).groupby("Member_ID").agg(
            First_Claim=("Treatment_Date", "min"),
            Total_Spent=("Amount_Paid", "sum"),
        )
        # Merge
        ehc = member_join.join(member_claims, how="inner")
        ehc["Days"] = (ehc["First_Claim"] - ehc["Join_Date"]).dt.days
        # Filter: spent over 200K within 30 days of enrolment
        ehc_flagged = ehc[(ehc["Days"].abs() <= 30) & (ehc["Total_Spent"] > 200_000)].sort_values("Total_Spent", ascending=False)
        early_claimers_count = len(ehc_flagged)

        # Get principal member name for initials
        principal_map = {}
        if "Principal_Member" in claims.columns:
            principal_map = claims.dropna(subset=["Member_ID", "Principal_Member"]).drop_duplicates("Member_ID").set_index("Member_ID")["Principal_Member"].to_dict()

        for mid, row in ehc_flagged.head(20).iterrows():
            full_name = str(principal_map.get(mid, "")).strip()
            initials = to_initials(full_name) if full_name else ""
            join_dt = row["Join_Date"].strftime("%Y-%m-%d") if pd.notna(row["Join_Date"]) else ""
            first_dt = row["First_Claim"].strftime("%Y-%m-%d") if pd.notna(row["First_Claim"]) else ""
            days_val = int(row["Days"])
            days_str = f"{days_val}d"
            premium_val = row["Premium"] if "Premium" in prod.columns and pd.notna(row["Premium"]) and row["Premium"] > 0 else 0
            spend_ratio = row["Total_Spent"] / premium_val if premium_val > 0 else 0
            ratio_cls = "danger" if spend_ratio > 5 else ("warning" if spend_ratio > 2 else "")
            ratio_str = f"{spend_ratio:.1f}x" if premium_val > 0 else "N/A"
            early_claimers_rows += f'<tr><td>{mid}</td><td>{initials}</td><td class="num">{join_dt}</td><td class="num">{first_dt}</td><td class="num">{days_str}</td><td class="num">{fmt_full(row["Total_Spent"])}</td><td class="num">{fmt_full(premium_val) if premium_val > 0 else "—"}</td><td class="num {ratio_cls}">{ratio_str}</td></tr>'

    # ── What Went Wrong (MLR > 90%) ──
    what_went_wrong_html = ""
    if mlr_pct > 90 and earned_total > 0:
        issues = []
        # Identify top cost-driving disease categories
        if "Diagnosis_Description" in claims.columns and "Disease_Category" in claims.columns:
            top_cats = claims.groupby("Disease_Category")["Amount_Paid"].sum().sort_values(ascending=False).head(3)
            top_cat_names = ", ".join(top_cats.index.tolist())
            top_cat_pct = (top_cats.sum() / paid_total * 100)
            issues.append(f"The top 3 disease categories (<strong>{top_cat_names}</strong>) account for <strong>{top_cat_pct:.0f}%</strong> of total paid claims, indicating heavy cost concentration.")
        # High-cost providers
        if "Provider" in claims.columns:
            top3_prov = claims.groupby("Provider")["Amount_Paid"].sum().sort_values(ascending=False).head(3)
            top3_prov_pct = (top3_prov.sum() / paid_total * 100)
            issues.append(f"The top 3 providers consume <strong>{top3_prov_pct:.0f}%</strong> of paid claims. Provider concentration risk is elevated.")
        # High utilisation per member
        if unique_members > 0:
            freq = unique_claims / unique_members
            issues.append(f"Average claims frequency is <strong>{freq:.1f} visits per member</strong>. {'This is high and suggests over-utilisation.' if freq > 3 else 'Monitor for upward trends.'}")
        # Pipeline pressure
        if pipeline_total > 0 and earned_total > 0:
            pipeline_pct = pipeline_total / earned_total * 100
            if pipeline_pct > 10:
                issues.append(f"Outstanding pipeline claims represent <strong>{pipeline_pct:.1f}%</strong> of earned premium — a significant liability exposure.")
        # IBNR warning
        if ibnr_total > 0 and earned_total > 0:
            ibnr_pct_ep = ibnr_total / earned_total * 100
            if ibnr_pct_ep > 5:
                issues.append(f"IBNR reserve estimate is <strong>{ibnr_pct_ep:.1f}%</strong> of earned premium, indicating delayed claims reporting.")

        if issues:
            items = "".join(f"<li>{i}</li>" for i in issues)
            what_went_wrong_html = f'''<div class="alert-card warning-card">
    <h2>&#9888; What Went <span class="accent">Wrong</span></h2>
    <p class="alert-subtitle">MLR is at <strong>{pct(mlr_pct)}</strong> — claims are consuming a disproportionate share of premium income.</p>
    <ul class="insight-list">{items}</ul></div>'''

    # ── Is the Plan Adequately Priced? (MLR > 100%) ──
    pricing_html = ""
    if mlr_pct > 100 and earned_total > 0:
        shortfall = total_incurred - earned_total
        required_premium = total_incurred / 0.85  # target 85% MLR
        premium_increase_pct = ((required_premium - earned_total) / earned_total * 100) if earned_total > 0 else 0
        per_member_deficit = shortfall / total_enrolled if total_enrolled > 0 else 0

        pricing_points = []
        pricing_points.append(f"Total incurred claims (<strong>{fmt_full(total_incurred)}</strong>) exceed earned premium (<strong>{fmt_full(earned_total)}</strong>) by <strong>{fmt_full(shortfall)}</strong>.")
        pricing_points.append(f"To achieve a target MLR of 85%, the required annual premium would be approximately <strong>{fmt_full(required_premium)}</strong> — a <strong>{premium_increase_pct:.0f}%</strong> increase.")
        if per_member_deficit > 0:
            pricing_points.append(f"The per-member deficit is <strong>{fmt_full(per_member_deficit)}</strong>, which must be recovered through premium adjustment or benefit redesign.")
        pricing_points.append("The current premium structure is <strong>not sustainable</strong>. Without corrective action, the plan will continue to operate at a loss.")

        items = "".join(f"<li>{i}</li>" for i in pricing_points)
        pricing_html = f'''<div class="alert-card danger-card">
    <h2>&#128200; Is the Plan <span class="accent">Adequately Priced?</span></h2>
    <p class="alert-subtitle">MLR exceeds 100% — the plan is paying out more in claims than it earns in premium.</p>
    <ul class="insight-list">{items}</ul></div>'''

    # ── Key Recommendations ──
    recommendations = []
    if mlr_pct > 100:
        recommendations.append("Initiate an immediate premium review. Current pricing does not cover incurred claims and the plan is operating at a loss.")
    elif mlr_pct > 90:
        recommendations.append("Conduct a premium adequacy assessment at the next renewal cycle. The MLR is approaching unsustainable levels.")
    if "Provider" in claims.columns:
        recommendations.append("Negotiate provider tariff agreements and expand the preferred provider network to control unit costs.")
    if "Disease_Category" in claims.columns:
        top_cat = claims.groupby("Disease_Category")["Amount_Paid"].sum().idxmax()
        recommendations.append(f"Implement targeted wellness and disease management programmes for <strong>{top_cat}</strong>, the highest cost disease category.")
    if unique_members > 0 and unique_claims / unique_members > 3:
        recommendations.append("Introduce utilisation management controls such as pre-authorisation for high-cost procedures and specialist referrals.")
    if pipeline_total > 0 and earned_total > 0 and (pipeline_total / earned_total * 100) > 10:
        recommendations.append("Accelerate claims adjudication to reduce pipeline backlog and improve reserve forecasting accuracy.")
    if ibnr_total > 0:
        recommendations.append("Strengthen claims reporting timelines with providers to reduce IBNR exposure and improve loss estimation.")
    recommendations.append("Review benefit design — consider co-payments, sub-limits, or exclusions on high-frequency, low-severity claims to reduce overall utilisation.")
    recommendations.append("Schedule quarterly claims experience reviews to monitor trends and enable early intervention.")

    reco_items = "".join(f"<li>{r}</li>" for r in recommendations)
    recommendations_html = f'''<div class="alert-card reco-card">
    <h2>&#128161; Key <span class="accent">Recommendations</span></h2>
    <ol class="insight-list numbered">{reco_items}</ol></div>'''

    logo = get_logo_b64()
    clean_name = str(client_name).strip().replace(" ", "_")[:30]

    # ── Additional info ──
    fee_note = ""
    if broker_fee > 0:
        fee_note = f"<tr><td>Broker Fee</td><td>{broker_fee}%</td></tr>"
    addl_note = f'<p style="color:var(--text-muted);font-size:12px;margin-top:8px;">{additional_reqs}</p>' if additional_reqs else ""

    html = f"""<!DOCTYPE html>
<html lang="en">
<head>
<meta charset="utf-8">
<meta name="viewport" content="width=device-width, initial-scale=1">
<title>{client_name} — Analytics Report</title>
<style>
  @import url('https://fonts.googleapis.com/css2?family=Inter:wght@300;400;500;600;700;800;900&display=swap');
  :root {{ --navy: #1A1A2E; --crimson: #C8102E; --coral: #E87722; --cream: #FAF7F2; --light-blue: #E8F4FD; --light-grey: #F4F4F6; --medium-grey: #E8E8EC; --text-dark: #1A1A2E; --text-muted: #6B7280; --white: #FFFFFF; }}
  * {{ margin: 0; padding: 0; box-sizing: border-box; }}
  body {{ font-family: 'Inter', sans-serif; background: var(--cream); color: var(--text-dark); line-height: 1.6; font-size: 14px; }}
  .container {{ max-width: 1140px; margin: 0 auto; padding: 40px 20px; }}
  .header {{ background: var(--navy); color: var(--white); padding: 40px; border-radius: 16px; margin-bottom: 30px; display: flex; align-items: center; gap: 24px; }}
  .header img {{ height: 60px; }}
  .header h1 {{ font-weight: 800; font-size: 28px; }}
  .header .subtitle {{ font-size: 14px; opacity: 0.7; }}
  .section {{ background: var(--white); border-radius: 14px; padding: 32px; margin-bottom: 24px; box-shadow: 0 1px 3px rgba(0,0,0,0.04); }}
  .section h2 {{ font-weight: 700; font-size: 20px; color: var(--navy); margin-bottom: 20px; padding-bottom: 10px; border-bottom: 2px solid var(--medium-grey); }}
  .section h2 span.accent {{ color: var(--crimson); }}
  .kpi-grid {{ display: grid; grid-template-columns: repeat(auto-fit, minmax(200px, 1fr)); gap: 16px; }}
  .kpi {{ background: var(--white); border: 1px solid var(--medium-grey); border-radius: 12px; padding: 22px 18px; text-align: center; }}
  .kpi .label {{ font-size: 11px; font-weight: 600; text-transform: uppercase; letter-spacing: 0.8px; color: var(--text-muted); margin-bottom: 8px; }}
  .kpi .value {{ font-weight: 800; font-size: 22px; color: var(--navy); }}
  .kpi.highlight {{ border-color: var(--crimson); border-width: 2px; }}
  .kpi.highlight .value {{ color: var(--crimson); }}
  .data-table {{ width: 100%; border-collapse: collapse; font-size: 13px; }}
  .data-table thead th {{ background: var(--navy); color: var(--white); padding: 12px 10px; text-align: left; font-weight: 600; font-size: 11px; text-transform: uppercase; }}
  .data-table thead th.num {{ text-align: right; }}
  .data-table tbody td {{ padding: 10px; border-bottom: 1px solid var(--light-grey); }}
  .data-table tbody td.num {{ text-align: right; }}
  .data-table tbody tr:hover {{ background: var(--light-blue); }}
  .data-table tbody tr:nth-child(even) {{ background: var(--light-grey); }}
  .data-table tbody td.good {{ color: #16a34a; font-weight: 700; }}
  .data-table tbody td.warning {{ color: var(--coral); font-weight: 700; }}
  .data-table tbody td.danger {{ color: var(--crimson); font-weight: 700; }}
  .mlr-card {{ background: var(--navy); color: var(--white); border-radius: 14px; padding: 32px; margin-bottom: 24px; }}
  .mlr-card h2 {{ color: var(--white); border: none; }}
  .mlr-table {{ width: 100%; border-collapse: collapse; }}
  .mlr-table td {{ padding: 10px 12px; border-bottom: 1px solid rgba(255,255,255,0.08); font-size: 13px; }}
  .mlr-table td:last-child {{ text-align: right; font-weight: 600; }}
  .mlr-table tr.total td {{ border-top: 2px solid var(--crimson); font-weight: 800; font-size: 15px; }}
  .mlr-table tr.total td:last-child {{ color: var(--coral); }}
  .footer {{ text-align: center; color: var(--text-muted); font-size: 11px; margin-top: 20px; padding: 20px; }}
  .dl-btn {{ display: inline-flex; align-items: center; gap: 6px; padding: 10px 20px; background: var(--crimson); color: var(--white); border: none; border-radius: 8px; font-family: inherit; font-size: 13px; font-weight: 700; text-decoration: none; cursor: pointer; transition: background 0.2s; margin-left: auto; }}
  .dl-btn:hover {{ background: #a00d24; }}
  .dl-btn svg {{ width: 16px; height: 16px; }}
  .alert-card {{ border-radius: 14px; padding: 32px; margin-bottom: 24px; }}
  .alert-card h2 {{ font-weight: 700; font-size: 20px; margin-bottom: 12px; padding-bottom: 10px; border-bottom: 2px solid rgba(0,0,0,0.08); }}
  .alert-card h2 span.accent {{ color: var(--crimson); }}
  .alert-subtitle {{ font-size: 14px; color: var(--text-muted); margin-bottom: 16px; font-weight: 500; }}
  .alert-subtitle strong {{ color: var(--text-dark); }}
  .insight-list {{ margin: 0; padding-left: 20px; }}
  .insight-list li {{ padding: 8px 0; font-size: 13px; line-height: 1.7; border-bottom: 1px solid rgba(0,0,0,0.04); }}
  .insight-list li:last-child {{ border: none; }}
  .insight-list li strong {{ color: var(--navy); }}
  .warning-card {{ background: #FFF7ED; border: 2px solid var(--coral); box-shadow: 0 2px 8px rgba(232,119,34,0.1); }}
  .warning-card h2 {{ color: var(--coral); }}
  .danger-card {{ background: #FEF2F2; border: 2px solid var(--crimson); box-shadow: 0 2px 8px rgba(200,16,46,0.1); }}
  .danger-card h2 {{ color: var(--crimson); }}
  .reco-card {{ background: #F0FDF4; border: 2px solid #16a34a; box-shadow: 0 2px 8px rgba(22,163,74,0.1); }}
  .reco-card h2 {{ color: #16a34a; }}
  .reco-card h2 span.accent {{ color: #15803d; }}
  .insight-list.numbered {{ list-style-type: decimal; }}
  @media print {{ .dl-btn {{ display: none; }} }}
</style>
</head>
<body>
<div class="container">
  <div class="header">
    {'<img src="data:image/jpeg;base64,' + logo + '" alt="Leadway Health">' if logo else ''}
    <div><h1>{client_name}</h1><div class="subtitle">Analytics Report &nbsp;|&nbsp; {pd.Timestamp.now().strftime('%d %B %Y')}</div></div>
    <button class="dl-btn" onclick="downloadReport()" style="margin-left:auto;"><svg xmlns='http://www.w3.org/2000/svg' fill='none' viewBox='0 0 24 24' stroke='currentColor' stroke-width='2'><path stroke-linecap='round' stroke-linejoin='round' d='M4 16v2a2 2 0 002 2h12a2 2 0 002-2v-2M7 10l5 5m0 0l5-5m-5 5V3'/></svg>Download</button>
  </div>
  {addl_note}
  <div class="section">
    <h2>Executive <span class="accent">Overview</span></h2>
    <div class="kpi-grid">
      <div class="kpi highlight"><div class="label">Total Incurred</div><div class="value">{fmt(total_incurred)}</div></div>
      <div class="kpi"><div class="label">Paid Claims</div><div class="value">{fmt(paid_total)}</div></div>
      <div class="kpi"><div class="label">Pipeline</div><div class="value">{fmt(pipeline_total)}</div></div>
      <div class="kpi"><div class="label">Earned Premium</div><div class="value">{fmt(earned_total)}</div></div>
      <div class="kpi"><div class="label">Written Premium</div><div class="value">{fmt(written_total)}</div></div>
      <div class="kpi highlight"><div class="label">MLR</div><div class="value">{pct(mlr_pct)}</div></div>
      <div class="kpi highlight"><div class="label">COR</div><div class="value">{pct(cor_pct)}</div></div>
      <div class="kpi"><div class="label">Members</div><div class="value">{unique_members:,}</div></div>
      <div class="kpi"><div class="label">Unique Claims</div><div class="value">{unique_claims:,}</div></div>
      <div class="kpi"><div class="label">Avg/Member</div><div class="value">{fmt(avg_per_member)}</div></div>
    </div>
  </div>
  <div class="mlr-card">
    <h2>MLR &amp; COR</h2>
    <table class="mlr-table">
      <tr><td>Paid Claims</td><td>{fmt_full(paid_total)}</td></tr>
      <tr><td>Pipeline Claims</td><td>{fmt_full(pipeline_total)}</td></tr>
      <tr><td>IBNR</td><td>{fmt_full(ibnr_total)}</td></tr>
      <tr class="total"><td>Total Incurred</td><td>{fmt_full(total_incurred)}</td></tr>
      <tr><td style="padding-top:16px;">Earned Premium</td><td style="padding-top:16px;">{fmt_full(earned_total)}</td></tr>
      <tr class="total"><td>MLR</td><td>{pct(mlr_pct)}</td></tr>
      <tr><td style="padding-top:16px;">Admin ({admin_pct}%)</td><td style="padding-top:16px;">{fmt_full(admin_amount)}</td></tr>
      <tr><td>NHIA (2%)</td><td>{fmt_full(nhia_amount)}</td></tr>
      {fee_note}
      <tr class="total"><td>COR</td><td>{pct(cor_pct)}</td></tr>
    </table>
  </div>
  <div class="section"><h2>Monthly Claims <span class="accent">Trend</span></h2>
    <table class="data-table"><thead><tr><th>Month</th><th class="num">Unique Claims</th><th class="num">Members</th><th class="num">Total Paid</th><th class="num">Avg/Member</th><th class="num">Visits/Member</th></tr></thead>
    <tbody>{monthly_rows}</tbody></table></div>
  <div class="section"><h2>Top <span class="accent">Providers</span></h2>
    <table class="data-table"><thead><tr><th>#</th><th>Provider</th><th class="num">Total Paid</th><th class="num">%</th><th class="num">Claims</th></tr></thead>
    <tbody>{prov_rows}</tbody></table></div>
  <div class="section"><h2>Department <span class="accent">Breakdown</span></h2>
    <table class="data-table"><thead><tr><th>#</th><th>Department</th><th class="num">Total Paid</th><th class="num">%</th><th class="num">Claims</th></tr></thead>
    <tbody>{dept_rows}</tbody></table></div>
  {"" if not disease_cat_rows else '<div class="section"><h2>Top 10 <span class="accent">Disease Categories</span></h2><table class="data-table"><thead><tr><th>#</th><th>Disease Category</th><th class="num">Total Paid</th><th class="num">%</th><th class="num">Claims</th></tr></thead><tbody>' + disease_cat_rows + '</tbody></table></div>'}
  {"" if not age_mlr_rows else '<div class="section"><h2>MLR by <span class="accent">Age Group</span></h2><table class="data-table"><thead><tr><th>Age Group</th><th class="num">Members</th><th class="num">Claims</th><th class="num">Total Paid</th><th class="num">Earned Premium</th><th class="num">MLR</th><th class="num">Avg/Member</th></tr></thead><tbody>' + age_mlr_rows + '</tbody></table></div>'}
  {"" if not early_claimers_rows else '<div class="section"><h2>Early High <span class="accent">Claimers</span></h2><p style="color:var(--text-muted);font-size:13px;margin-bottom:16px;">Members who spent over &#8358;200K within 30 days of enrolment (' + str(early_claimers_count) + ' flagged)</p><table class="data-table"><thead><tr><th>Member ID</th><th>Principal</th><th class="num">Join Date</th><th class="num">First Claim</th><th class="num">Days</th><th class="num">Total Spent</th><th class="num">Premium</th><th class="num">Spend/Premium</th></tr></thead><tbody>' + early_claimers_rows + '</tbody></table></div>'}
  {what_went_wrong_html}
  {pricing_html}
  {recommendations_html}
  <div class="footer">Generated by Leadway Health Analytics &nbsp;|&nbsp; {pd.Timestamp.now().strftime('%d %B %Y')}</div>
</div>
<script>
function downloadReport() {{
  var el = document.createElement('a');
  el.setAttribute('href', 'data:text/html;charset=utf-8,' + encodeURIComponent(document.documentElement.outerHTML));
  el.setAttribute('download', '{clean_name}_Report.html');
  el.style.display = 'none';
  document.body.appendChild(el);
  el.click();
  document.body.removeChild(el);
}}
</script>
</body></html>"""

    # Save report
    report_path = REPORTS_DIR / f"{clean_name}_Report_{session_id}.html"
    report_path.write_text(html)
    return report_path


# ═══════════════════════════════════════════════
# PROVIDER INTELLIGENCE — Submodules
# ═══════════════════════════════════════════════

PROVIDER_MODULES = {
    "tariff-intelligence": {"title": "Tariff Intelligence", "desc": "Analyse provider charges against master tariff, identify outliers and overcharges.", "icon": "&#128200;", "color": "#1B1464"},
    "fwa-insights": {"title": "FWA Insights", "desc": "Fraud, waste & abuse detection — flag suspicious patterns, duplicate claims, upcoding.", "icon": "&#128270;", "color": "#1B1464"},
    "tariff-mapper": {"title": "Tariff Mapper", "desc": "Map provider codes to master tariff, reconcile naming differences, standardise billing.", "icon": "&#128268;", "color": "#1B1464"},
    "plan-access": {"title": "Plan Access Argument Generator", "desc": "Generate data-backed arguments for provider tier upgrades or plan access negotiations.", "icon": "&#128220;", "color": "#1B1464"},
}


@app.route("/tariff-intelligence")
@app.route("/fwa-insights")
@app.route("/tariff-mapper")
@app.route("/plan-access")
@login_required
def provider_submodule():
    slug = request.path.strip("/")
    mod = PROVIDER_MODULES.get(slug)
    if not mod:
        return redirect(url_for("index"))
    logo = get_logo_b64()
    return render_template("provider_module.html", logo=logo, module=mod, slug=slug)


@app.route("/tariff-intelligence/upload", methods=["POST"])
@app.route("/fwa-insights/upload", methods=["POST"])
@app.route("/tariff-mapper/upload", methods=["POST"])
@app.route("/plan-access/upload", methods=["POST"])
@login_required
def provider_submodule_upload():
    slug = request.path.replace("/upload", "").strip("/")
    flash(f"{PROVIDER_MODULES.get(slug, {}).get('title', 'Module')} — coming soon.")
    return redirect(f"/{slug}")


# ═══════════════════════════════════════════════
# PROVIDER ANALYTICS — Full module
# ═══════════════════════════════════════════════

@app.route("/provider-analytics")
@login_required
def provider_analytics():
    logo = get_logo_b64()
    report_list = list_reports("ProviderAnalytics_")
    return render_template("provider_analytics.html", logo=logo, reports=report_list)


@app.route("/provider-analytics/upload", methods=["POST"])
@login_required
def provider_analytics_upload():
    import pandas as pd
    import numpy as np

    claims_file = request.files.get("claims_file")
    if not claims_file or claims_file.filename == "":
        flash("Please upload a claims file.")
        return redirect(url_for("provider_analytics"))

    session_id = str(uuid.uuid4())[:8]
    upload_dir = app.config["UPLOAD_FOLDER"] / session_id
    upload_dir.mkdir(parents=True, exist_ok=True)
    fpath = upload_dir / claims_file.filename
    claims_file.save(fpath)

    # Parse
    if str(fpath).endswith(".csv"):
        df = pd.read_csv(fpath)
    else:
        df = pd.read_excel(fpath)
    df.columns = [c.replace("\n", " ").replace("\r", " ").strip() for c in df.columns]

    # ── Column mapping ──
    col_map = {}
    for col in df.columns:
        cl = col.lower().strip()
        if cl in ("amt paid", "claims paid", "amount paid"): col_map[col] = "Amt_Paid"
        elif cl in ("amt claimed", "amount claimed"): col_map[col] = "Amt_Claimed"
        elif cl in ("claim number", "claim no", "claim numb"): col_map[col] = "Claim_No"
        elif cl in ("membershipno", "enrolee id", "member id", "member enrollee id"): col_map[col] = "Enrolee_ID"
        elif cl in ("treatment date", "encounter date"): col_map[col] = "Treatment_Date"
        elif cl in ("provider", "provider name"): col_map[col] = "Provider"
        elif cl in ("group name", "group"): col_map[col] = "Group_Name"
        elif cl in ("scheme", "scheme name"): col_map[col] = "Scheme"
        elif cl in ("service type",): col_map[col] = "Service_Type"
        elif cl in ("tariff descr", "tariff description"): col_map[col] = "Tariff_Descr"
        elif cl in ("benefit", "benefit type"): col_map[col] = "Benefit"
        elif cl in ("diag descr", "diagnosis description"): col_map[col] = "Diag_Descr"
        elif cl in ("diagnosis",): col_map[col] = "Diagnosis"
        elif cl in ("claim status",): col_map[col] = "Claim_Status"
        elif cl in ("relationship type", "relationship"): col_map[col] = "Relationship"
        elif cl in ("member gender", "member sex", "gender"): col_map[col] = "Gender"
        elif cl in ("member age", "age"): col_map[col] = "Member_Age"
        elif cl in ("prov location", "provider location", "state"): col_map[col] = "Prov_Location"
    df.rename(columns=col_map, inplace=True)

    # ── Parse types ──
    for col in ["Amt_Paid", "Amt_Claimed"]:
        if col in df.columns:
            df[col] = pd.to_numeric(df[col].astype(str).str.replace(",", "").str.replace("₦", ""), errors="coerce").fillna(0)

    if "Treatment_Date" in df.columns:
        raw = df["Treatment_Date"].astype(str).str.strip().replace({"############": pd.NA, "nan": pd.NA})
        num = pd.to_numeric(raw, errors="coerce")
        excel_like = num.notna() & (num > 30000) & (num < 60000)
        if excel_like.sum() > num.notna().sum() * 0.3 and num.notna().sum() > 0:
            df["Treatment_Date"] = pd.NaT
            df.loc[excel_like, "Treatment_Date"] = pd.to_datetime("1899-12-30") + pd.to_timedelta(num[excel_like], unit="D")
        else:
            df["Treatment_Date"] = pd.to_datetime(raw, errors="coerce", dayfirst=True)

    # Normalise service type
    if "Service_Type" in df.columns:
        st = df["Service_Type"].astype(str).str.strip().str.lower()
        df.loc[st.isin(["outpatient", "opd", "out-patient", "mtn pha"]), "Service_Type"] = "Outpatient"
        df.loc[st.isin(["inpatient", "ipd", "in-patient"]), "Service_Type"] = "Inpatient"

    # Exclude abandoned
    if "Claim_Status" in df.columns:
        df = df[~df["Claim_Status"].str.lower().str.strip().isin(["abandoned", "rejected", "declined"])]

    # Effective spend
    df["Spend"] = df["Amt_Paid"] if "Amt_Paid" in df.columns else 0
    if "Amt_Claimed" in df.columns and "Amt_Paid" in df.columns:
        zero_paid = df["Amt_Paid"].fillna(0) == 0
        df.loc[zero_paid, "Spend"] = df.loc[zero_paid, "Amt_Claimed"]

    # Family ID
    if "Enrolee_ID" in df.columns:
        df["Family_ID"] = df["Enrolee_ID"].astype(str).str[:8]

    # ── Generate report ──
    try:
        report_path = generate_provider_report(df, session_id)
        return redirect(url_for("view_report", filename=report_path.name))
    except Exception:
        import traceback
        return f"<pre>Error generating report:\n{traceback.format_exc()}</pre>", 500


def generate_provider_report(df, session_id):
    """Generate a comprehensive Provider Analytics HTML report."""
    import pandas as pd
    import numpy as np

    REPORTS_DIR.mkdir(parents=True, exist_ok=True)
    logo = get_logo_b64()

    def fmt(x):
        if abs(x) >= 1_000_000: return f"&#8358;{x/1_000_000:,.2f}M"
        if abs(x) >= 1_000: return f"&#8358;{x/1_000:,.0f}K"
        return f"&#8358;{x:,.0f}"

    def fmt_full(x): return f"&#8358;{x:,.0f}"
    def pct(x): return f"{x:.1f}%"

    total_spend = float(df["Spend"].sum())
    total_claims = int(df["Claim_No"].nunique()) if "Claim_No" in df.columns else len(df)
    total_members = int(df["Enrolee_ID"].nunique()) if "Enrolee_ID" in df.columns else 0
    providers_list = sorted(df["Provider"].dropna().unique().tolist()) if "Provider" in df.columns else []
    provider_label = providers_list[0] if len(providers_list) == 1 else f"{len(providers_list)} Providers"

    # ── 1. Monthly Overview ──
    monthly_rows = ""
    if "Treatment_Date" in df.columns:
        d = df.dropna(subset=["Treatment_Date"]).copy()
        d["Month"] = d["Treatment_Date"].dt.to_period("M")
        g = d.groupby("Month").agg(spend=("Spend", "sum"),
            members=("Enrolee_ID", "nunique") if "Enrolee_ID" in df.columns else ("Spend", "count"),
            visits=("Claim_No", "nunique") if "Claim_No" in df.columns else ("Spend", "count")).sort_index()
        prev = None
        for p, r in g.iterrows():
            s, m, v = float(r["spend"]), int(r["members"]), int(r["visits"])
            apm = round(s / max(m, 1))
            apv = round(s / max(v, 1))
            mom = f'{((s - prev) / prev * 100):+.1f}%' if prev and prev > 0 else "—"
            monthly_rows += f'<tr><td>{p.to_timestamp().strftime("%b %Y")}</td><td class="num">{fmt_full(s)}</td><td class="num">{m:,}</td><td class="num">{v:,}</td><td class="num">{fmt_full(apm)}</td><td class="num">{fmt_full(apv)}</td><td class="num">{mom}</td></tr>'
            prev = s

    # ── 2. Service Type Split ──
    svc_rows = ""
    if "Service_Type" in df.columns:
        for st, grp in df.groupby("Service_Type"):
            s = float(grp["Spend"].sum())
            v = int(grp["Claim_No"].nunique()) if "Claim_No" in grp.columns else len(grp)
            svc_rows += f'<tr><td>{st}</td><td class="num">{fmt_full(s)}</td><td class="num">{pct(s / max(total_spend, 1) * 100)}</td><td class="num">{v:,}</td></tr>'

    # ── 3. Top Providers ──
    prov_rows = ""
    if "Provider" in df.columns:
        g = df.groupby("Provider").agg(spend=("Spend", "sum"),
            members=("Enrolee_ID", "nunique") if "Enrolee_ID" in df.columns else ("Spend", "count"),
            visits=("Claim_No", "nunique") if "Claim_No" in df.columns else ("Spend", "count")).sort_values("spend", ascending=False)
        for i, (name, r) in enumerate(g.head(20).iterrows(), 1):
            s = float(r["spend"])
            prov_rows += f'<tr><td>{i}</td><td>{name}</td><td class="num">{fmt_full(s)}</td><td class="num">{pct(s / max(total_spend, 1) * 100)}</td><td class="num">{int(r["members"]):,}</td><td class="num">{int(r["visits"]):,}</td><td class="num">{fmt_full(round(s / max(int(r["members"]), 1)))}</td></tr>'

    # ── 4. Groups ──
    group_rows = ""
    if "Group_Name" in df.columns:
        g = df.groupby("Group_Name").agg(spend=("Spend", "sum"),
            members=("Enrolee_ID", "nunique") if "Enrolee_ID" in df.columns else ("Spend", "count"),
            visits=("Claim_No", "nunique") if "Claim_No" in df.columns else ("Spend", "count")).sort_values("spend", ascending=False)
        for name, r in g.iterrows():
            s = float(r["spend"])
            group_rows += f'<tr><td>{name}</td><td class="num">{fmt_full(s)}</td><td class="num">{pct(s / max(total_spend, 1) * 100)}</td><td class="num">{int(r["members"]):,}</td><td class="num">{int(r["visits"]):,}</td><td class="num">{fmt_full(round(s / max(int(r["members"]), 1)))}</td></tr>'

    # ── 5. Top 30 Tariff Lines ──
    tariff_rows = ""
    if "Tariff_Descr" in df.columns:
        g = df.groupby("Tariff_Descr").agg(spend=("Spend", "sum"),
            members=("Enrolee_ID", "nunique") if "Enrolee_ID" in df.columns else ("Spend", "count"),
            utilized=("Claim_No", "nunique") if "Claim_No" in df.columns else ("Spend", "count")).sort_values("spend", ascending=False).head(30)
        cum = 0
        for i, (name, r) in enumerate(g.iterrows(), 1):
            s = float(r["spend"])
            p = s / max(total_spend, 1) * 100
            cum += p
            avg = round(s / max(int(r["utilized"]), 1))
            tariff_rows += f'<tr><td>{i}</td><td>{name}</td><td class="num">{fmt_full(s)}</td><td class="num">{pct(p)}</td><td class="num">{pct(cum)}</td><td class="num">{int(r["utilized"]):,}</td><td class="num">{fmt_full(avg)}</td></tr>'

    # ── 6. Chronic Medication ──
    chronic_rows = ""
    chronic_spend = 0
    chronic_members = 0
    if "Benefit" in df.columns:
        ch = df[df["Benefit"].str.contains("chronic", case=False, na=False)]
        chronic_spend = float(ch["Spend"].sum())
        chronic_members = int(ch["Enrolee_ID"].nunique()) if "Enrolee_ID" in ch.columns else 0
        drug_col = "Tariff_Descr" if "Tariff_Descr" in ch.columns else None
        if drug_col and not ch.empty:
            g = ch.groupby(drug_col).agg(spend=("Spend", "sum"),
                members=("Enrolee_ID", "nunique") if "Enrolee_ID" in ch.columns else ("Spend", "count"),
                dispensed=("Claim_No", "nunique") if "Claim_No" in ch.columns else ("Spend", "count")).sort_values("spend", ascending=False)
            for name, r in g.iterrows():
                s = float(r["spend"])
                chronic_rows += f'<tr><td>{name}</td><td class="num">{fmt_full(s)}</td><td class="num">{int(r["members"]):,}</td><td class="num">{int(r["dispensed"]):,}</td><td class="num">{fmt_full(round(s / max(int(r["dispensed"]), 1)))}</td></tr>'

    # ── 7. Diagnosis Patterns ──
    diag_rows = ""
    vague_keywords = ["unspecified", "nos", "other", "unknown"]
    diag_col = "Diag_Descr" if "Diag_Descr" in df.columns else ("Diagnosis" if "Diagnosis" in df.columns else None)
    total_vague = 0
    if diag_col:
        g = df.groupby(diag_col).agg(spend=("Spend", "sum"),
            members=("Enrolee_ID", "nunique") if "Enrolee_ID" in df.columns else ("Spend", "count"),
            visits=("Claim_No", "nunique") if "Claim_No" in df.columns else ("Spend", "count")).sort_values("spend", ascending=False)
        for name, r in g.head(20).iterrows():
            s = float(r["spend"])
            is_vague = any(kw in str(name).lower() for kw in vague_keywords)
            if is_vague: total_vague += s
            vague_cls = ' class="warning"' if is_vague else ""
            diag_rows += f'<tr><td{vague_cls}>{name}{"  ⚠" if is_vague else ""}</td><td class="num">{fmt_full(s)}</td><td class="num">{pct(s / max(total_spend, 1) * 100)}</td><td class="num">{int(r["members"]):,}</td><td class="num">{int(r["visits"]):,}</td></tr>'

    # ── 8. Top Enrollees ──
    enrollee_rows = ""
    if "Enrolee_ID" in df.columns:
        agg = {"spend": ("Spend", "sum")}
        if "Claim_No" in df.columns: agg["visits"] = ("Claim_No", "nunique")
        if "Provider" in df.columns: agg["hospitals"] = ("Provider", "nunique")
        g = df.groupby("Enrolee_ID").agg(**agg).sort_values("spend", ascending=False)
        for i, (eid, r) in enumerate(g.head(20).iterrows(), 1):
            s = float(r["spend"])
            v = int(r.get("visits", 0))
            h = int(r.get("hospitals", 0))
            flag = " ⚠" if h > 5 else ""
            top = " 🔝" if i <= 10 else ""
            enrollee_rows += f'<tr><td>{eid}{top}</td><td>{str(eid)[:8]}</td><td class="num">{fmt_full(s)}</td><td class="num">{v:,}</td><td class="num">{h}{flag}</td></tr>'

    # ── Build HTML ──
    date_range = ""
    if "Treatment_Date" in df.columns:
        valid = df["Treatment_Date"].dropna()
        if len(valid) > 0:
            date_range = f'{valid.min().strftime("%b %Y")} — {valid.max().strftime("%b %Y")}'

    html = f"""<!DOCTYPE html>
<html lang="en">
<head>
<meta charset="utf-8"><meta name="viewport" content="width=device-width, initial-scale=1">
<title>Provider Analytics — {provider_label}</title>
<style>
  @import url('https://fonts.googleapis.com/css2?family=Inter:wght@300;400;500;600;700;800;900&display=swap');
  :root {{ --navy: #1B1464; --red: #C61531; --orange: #F15A24; --dark: #262626; --cream: #FAF7F2; --light: #F4F4F6; --grey: #E6E6E6; --white: #FFFFFF; --muted: #6B7280; --green: #16a34a; }}
  * {{ margin:0; padding:0; box-sizing:border-box; }}
  body {{ font-family:'Inter',sans-serif; background:var(--cream); color:var(--dark); line-height:1.6; font-size:13px; }}
  .container {{ max-width:1140px; margin:0 auto; padding:40px 20px; }}
  .header {{ background:var(--navy); color:white; padding:36px; border-radius:16px; margin-bottom:28px; display:flex; align-items:center; gap:24px; }}
  .header img {{ height:56px; }}
  .header h1 {{ font-weight:800; font-size:26px; }}
  .header .sub {{ font-size:13px; opacity:0.6; margin-top:4px; }}
  .kpi-grid {{ display:grid; grid-template-columns:repeat(auto-fit,minmax(180px,1fr)); gap:14px; margin-bottom:24px; }}
  .kpi {{ background:white; border-radius:12px; padding:20px 16px; text-align:center; box-shadow:0 1px 3px rgba(0,0,0,0.04); }}
  .kpi .label {{ font-size:10px; font-weight:700; text-transform:uppercase; letter-spacing:0.8px; color:var(--muted); margin-bottom:6px; }}
  .kpi .value {{ font-weight:800; font-size:20px; color:var(--dark); }}
  .kpi.highlight {{ border:2px solid var(--red); }}
  .kpi.highlight .value {{ color:var(--red); }}
  .section {{ background:white; border-radius:14px; padding:28px; margin-bottom:22px; box-shadow:0 1px 3px rgba(0,0,0,0.04); }}
  .section h2 {{ font-weight:700; font-size:18px; color:var(--dark); margin-bottom:16px; padding-bottom:8px; border-bottom:2px solid var(--grey); }}
  .section h2 span {{ color:var(--navy); }}
  .data-table {{ width:100%; border-collapse:collapse; font-size:12px; }}
  .data-table thead th {{ background:var(--dark); color:white; padding:10px 8px; text-align:left; font-weight:600; font-size:10px; text-transform:uppercase; }}
  .data-table thead th.num {{ text-align:right; }}
  .data-table tbody td {{ padding:8px; border-bottom:1px solid var(--light); }}
  .data-table tbody td.num {{ text-align:right; font-weight:500; }}
  .data-table tbody td.warning {{ color:var(--orange); font-weight:600; }}
  .data-table tbody tr:hover {{ background:var(--light); }}
  .data-table tbody tr:nth-child(even) {{ background:#FAFAFA; }}
  .alert {{ border-radius:10px; padding:16px 20px; margin-bottom:16px; font-size:13px; }}
  .alert.warn {{ background:#FFF7ED; border:2px solid var(--orange); color:var(--orange); }}
  .footer {{ text-align:center; color:var(--muted); font-size:10px; margin-top:20px; padding:16px; }}
  .dl-btn {{ display:inline-flex; align-items:center; gap:6px; padding:10px 20px; background:var(--navy); color:white; border:none; border-radius:8px; font-family:inherit; font-size:12px; font-weight:700; text-decoration:none; cursor:pointer; margin-left:auto; }}
  .dl-btn:hover {{ background:var(--dark); }}
  @media print {{ .dl-btn {{ display:none; }} }}
</style>
</head>
<body>
<div class="container">
  <div class="header">
    {'<img src="data:image/jpeg;base64,' + logo + '" alt="Leadway Health">' if logo else ''}
    <div>
      <h1>Provider Analytics</h1>
      <div class="sub">{provider_label} &nbsp;|&nbsp; {date_range} &nbsp;|&nbsp; {pd.Timestamp.now().strftime('%d %B %Y')}</div>
    </div>
    <button class="dl-btn" onclick="(function(){{var a=document.createElement('a');a.href='data:text/html;charset=utf-8,'+encodeURIComponent(document.documentElement.outerHTML);a.download='ProviderAnalytics_{session_id}.html';a.click();}})()">&#11015; Download</button>
  </div>

  <div class="kpi-grid">
    <div class="kpi highlight"><div class="label">Total Spend</div><div class="value">{fmt(total_spend)}</div></div>
    <div class="kpi"><div class="label">Unique Members</div><div class="value">{total_members:,}</div></div>
    <div class="kpi"><div class="label">Unique Claims</div><div class="value">{total_claims:,}</div></div>
    <div class="kpi"><div class="label">Avg / Visit</div><div class="value">{fmt(round(total_spend / max(total_claims, 1)))}</div></div>
    <div class="kpi"><div class="label">Avg / Member</div><div class="value">{fmt(round(total_spend / max(total_members, 1)))}</div></div>
    <div class="kpi"><div class="label">Providers</div><div class="value">{len(providers_list)}</div></div>
  </div>

  {"" if not svc_rows else '<div class="section"><h2>Service Type <span>Split</span></h2><table class="data-table"><thead><tr><th>Service Type</th><th class="num">Amount Paid</th><th class="num">% of Total</th><th class="num">Visits</th></tr></thead><tbody>' + svc_rows + '</tbody></table></div>'}

  {"" if not monthly_rows else '<div class="section"><h2>Monthly Claims <span>Trend</span></h2><table class="data-table"><thead><tr><th>Month</th><th class="num">Amount Paid</th><th class="num">Members</th><th class="num">Visits</th><th class="num">Avg/Member</th><th class="num">Avg/Visit</th><th class="num">MoM</th></tr></thead><tbody>' + monthly_rows + '</tbody></table></div>'}

  {"" if not prov_rows else '<div class="section"><h2>Provider <span>Summary</span></h2><table class="data-table"><thead><tr><th>#</th><th>Provider</th><th class="num">Amount Paid</th><th class="num">% Total</th><th class="num">Members</th><th class="num">Visits</th><th class="num">Per Member</th></tr></thead><tbody>' + prov_rows + '</tbody></table></div>'}

  {"" if not group_rows else '<div class="section"><h2>Group <span>Breakdown</span></h2><table class="data-table"><thead><tr><th>Group</th><th class="num">Amount Paid</th><th class="num">% Total</th><th class="num">Members</th><th class="num">Visits</th><th class="num">Per Member</th></tr></thead><tbody>' + group_rows + '</tbody></table></div>'}

  {"" if not tariff_rows else '<div class="section"><h2>Top 30 Tariff <span>Lines</span></h2><table class="data-table"><thead><tr><th>#</th><th>Service</th><th class="num">Amount Paid</th><th class="num">% Total</th><th class="num">Cum %</th><th class="num">Utilized</th><th class="num">Avg/Line</th></tr></thead><tbody>' + tariff_rows + '</tbody></table></div>'}

  {"" if chronic_spend == 0 else '<div class="section"><h2>Chronic <span>Medication</span></h2><div class="kpi-grid" style="margin-bottom:16px"><div class="kpi"><div class="label">Chronic Spend</div><div class="value">' + fmt(chronic_spend) + '</div></div><div class="kpi"><div class="label">Members on Chronic</div><div class="value">' + str(chronic_members) + '</div></div><div class="kpi"><div class="label">% of Total</div><div class="value">' + pct(chronic_spend / max(total_spend, 1) * 100) + '</div></div></div><table class="data-table"><thead><tr><th>Drug</th><th class="num">Total Spend</th><th class="num">Members</th><th class="num">Dispensed</th><th class="num">Avg/Dispense</th></tr></thead><tbody>' + chronic_rows + '</tbody></table></div>'}

  {"" if not diag_rows else ('<div class="alert warn">⚠ ' + pct(total_vague / max(total_spend, 1) * 100) + ' of spend is on vague/unspecified diagnoses — review for upcoding risk.</div>' if total_vague > total_spend * 0.15 else '') + '<div class="section"><h2>Diagnosis <span>Patterns</span></h2><table class="data-table"><thead><tr><th>Diagnosis</th><th class="num">Amount Paid</th><th class="num">% Total</th><th class="num">Members</th><th class="num">Visits</th></tr></thead><tbody>' + diag_rows + '</tbody></table></div>'}

  {"" if not enrollee_rows else '<div class="section"><h2>Top <span>Enrollees</span></h2><table class="data-table"><thead><tr><th>Enrolee ID</th><th>Family ID</th><th class="num">Total Paid</th><th class="num">Visits</th><th class="num">Hospitals</th></tr></thead><tbody>' + enrollee_rows + '</tbody></table></div>'}

  <div class="footer">Generated by Leadway Health Analytics — Provider Intelligence &nbsp;|&nbsp; {pd.Timestamp.now().strftime('%d %B %Y')}</div>
</div>
</body></html>"""

    report_path = REPORTS_DIR / f"ProviderAnalytics_{session_id}.html"
    report_path.write_text(html)
    return report_path


# ═══════════════════════════════════════════════
# MODULE 3 — Benefit Benchmarking
# ═══════════════════════════════════════════════

LEADWAY_BENEFITS = [
    {"category": "network", "row": "Provider Network", "PLUS": "Category D", "PRO": "Category C + D", "MAX": "Category B+C+D", "PROMAX": "Category A+B+C+D", "MAGNUM": "Category A+B+C+D", "MAGNUM PLUS": "Category A*+A+B+C+D"},
    {"category": "territory", "row": "Territory", "PLUS": "Nigeria + India", "PRO": "Nigeria + India", "MAX": "Nigeria + W.Africa + India + UAE", "PROMAX": "Local + Africa + India + UAE", "MAGNUM": "Local + Africa + India + UAE", "MAGNUM PLUS": "Local + Africa + India + UAE"},
    {"category": "admission", "row": "Admission / Ward Type", "PLUS": "General ward (Unlimited)", "PRO": "Semi Private (Unlimited)", "MAX": "Private ward (Unlimited)", "PROMAX": "Private ward (Unlimited)", "MAGNUM": "Private ward (Unlimited)", "MAGNUM PLUS": "Private ward (Unlimited)"},
    {"category": "icu", "row": "Intensive Care", "PLUS": "500,000", "PRO": "500,000", "MAX": "750,000", "PROMAX": "1,000,000", "MAGNUM": "1,500,000", "MAGNUM PLUS": "3,000,000"},
    {"category": "psychiatric", "row": "Psychiatric Hospitalisation", "PLUS": "2 days", "PRO": "3 days", "MAX": "5 days", "PROMAX": "7 days", "MAGNUM": "10 days", "MAGNUM PLUS": "15 days"},
    {"category": "maternity", "row": "Maternity (In-Network)", "PLUS": "Covered", "PRO": "Covered", "MAX": "Covered", "PROMAX": "Covered", "MAGNUM": "Covered", "MAGNUM PLUS": "Covered"},
    {"category": "neonatal", "row": "Neonatal Care (incl. incubator)", "PLUS": "500,000", "PRO": "750,000", "MAX": "3,500,000", "PROMAX": "5,000,000", "MAGNUM": "7,500,000", "MAGNUM PLUS": "10,000,000"},
    {"category": "congenital", "row": "Congenital Anomaly Treatment", "PLUS": "250,000", "PRO": "300,000", "MAX": "500,000", "PROMAX": "750,000", "MAGNUM": "1,000,000", "MAGNUM PLUS": "2,000,000"},
    {"category": "major disease", "row": "Major Disease (Cancer, Stroke)", "PLUS": "300,000", "PRO": "400,000", "MAX": "1,000,000", "PROMAX": "2,000,000", "MAGNUM": "5,000,000", "MAGNUM PLUS": "10,000,000"},
    {"category": "dialysis", "row": "Renal Dialysis", "PLUS": "100,000", "PRO": "4 sessions", "MAX": "6 sessions", "PROMAX": "8 sessions", "MAGNUM": "12 sessions", "MAGNUM PLUS": "Unlimited"},
    {"category": "surgery", "row": "Surgeries (Minor to Tertiary)", "PLUS": "400,000", "PRO": "500,000", "MAX": "Covered", "PROMAX": "Covered", "MAGNUM": "Covered", "MAGNUM PLUS": "Covered"},
    {"category": "spine", "row": "Spine Surgery", "PLUS": "300,000", "PRO": "400,000", "MAX": "1,000,000", "PROMAX": "1,500,000", "MAGNUM": "2,000,000", "MAGNUM PLUS": "3,000,000"},
    {"category": "optical", "row": "Lenses & Frames (Once/2 Years)", "PLUS": "20,000", "PRO": "25,000", "MAX": "45,000", "PROMAX": "60,000", "MAGNUM": "80,000", "MAGNUM PLUS": "100,000"},
    {"category": "dental", "row": "Primary Dental", "PLUS": "30,000", "PRO": "50,000", "MAX": "120,000", "PROMAX": "150,000", "MAGNUM": "200,000", "MAGNUM PLUS": "300,000"},
    {"category": "diagnostics", "row": "Advanced Diagnostics (MRI, CT)", "PLUS": "70,000", "PRO": "100,000", "MAX": "250,000", "PROMAX": "400,000", "MAGNUM": "Covered", "MAGNUM PLUS": "Covered"},
    {"category": "physiotherapy", "row": "Physiotherapy", "PLUS": "50,000", "PRO": "75,000", "MAX": "100,000", "PROMAX": "150,000", "MAGNUM": "250,000", "MAGNUM PLUS": "Covered"},
    {"category": "chronic", "row": "Chronic Disease Management", "PLUS": "400,000", "PRO": "500,000", "MAX": "750,000", "PROMAX": "1,000,000", "MAGNUM": "Unlimited", "MAGNUM PLUS": "Unlimited"},
    {"category": "devices", "row": "Medical Devices & Appliances", "PLUS": "50,000", "PRO": "75,000", "MAX": "100,000", "PROMAX": "150,000", "MAGNUM": "200,000", "MAGNUM PLUS": "300,000"},
    {"category": "mortuary", "row": "Mortuary Fee per Family", "PLUS": "250,000", "PRO": "300,000", "MAX": "400,000", "PROMAX": "500,000", "MAGNUM": "750,000", "MAGNUM PLUS": "1,000,000"},
    {"category": "evacuation", "row": "International Evacuation", "PLUS": "Not Covered", "PRO": "Not Covered", "MAX": "Not Covered", "PROMAX": "Not Covered", "MAGNUM": "$50,000", "MAGNUM PLUS": "$100,000"},
    {"category": "gym", "row": "Gym Services", "PLUS": "2 sessions/month", "PRO": "4 sessions/month", "MAX": "Twice a week", "PROMAX": "Twice a week", "MAGNUM": "Twice a week", "MAGNUM PLUS": "Incl. i-Fitness"},
    {"category": "premium", "row": "Individual Premium", "PLUS": "121,268", "PRO": "164,200", "MAX": "262,275", "PROMAX": "382,582", "MAGNUM": "734,250", "MAGNUM PLUS": "1,737,750"},
    {"category": "family premium", "row": "Family Premium", "PLUS": "576,023", "PRO": "779,950", "MAX": "1,245,806", "PROMAX": "1,817,265", "MAGNUM": "3,487,688", "MAGNUM PLUS": "8,254,313"},
]


def _compare_values(val_a, val_b):
    """Compare two benefit values. Returns 'a-better', 'b-better', or 'same'."""
    import re
    a, b = str(val_a).strip(), str(val_b).strip()
    al, bl = a.lower(), b.lower()
    if al == bl: return "same"
    if al in ("not covered", "—", "", "nan") and bl not in ("not covered", "—", "", "nan"): return "b-better"
    if bl in ("not covered", "—", "", "nan") and al not in ("not covered", "—", "", "nan"): return "a-better"
    if "unlimited" in bl and "unlimited" not in al: return "b-better"
    if "unlimited" in al and "unlimited" not in bl: return "a-better"
    nums_a = re.findall(r'[\d,]+', a.replace("₦", "").replace("NGN", "").replace("N", ""))
    nums_b = re.findall(r'[\d,]+', b.replace("₦", "").replace("NGN", "").replace("N", ""))
    if nums_a and nums_b:
        try:
            na = float(nums_a[0].replace(",", ""))
            nb = float(nums_b[0].replace(",", ""))
            if nb > na: return "b-better"
            if na > nb: return "a-better"
        except ValueError: pass
    return "same"


@app.route("/benefit-benchmarking")
@login_required
def benefit_benchmarking():
    logo = get_logo_b64()
    report_list = list_reports("Benefit_")
    return render_template("benefit_benchmarking.html", logo=logo, reports=report_list)


@app.route("/benefit-benchmarking/compare", methods=["POST"])
@login_required
def benefit_compare():
    import pandas as pd
    plan_a = request.form.get("plan_a", "PRO")
    plan_b = request.form.get("plan_b", "MAX")
    if plan_a == plan_b:
        flash("Please select two different plans.")
        return redirect(url_for("benefit_benchmarking"))

    logo = get_logo_b64()
    sid = str(uuid.uuid4())[:8]
    REPORTS_DIR.mkdir(parents=True, exist_ok=True)

    rows = ""
    for b in LEADWAY_BENEFITS:
        va, vb = b.get(plan_a, "—"), b.get(plan_b, "—")
        cmp = _compare_values(va, vb)
        ca = "better" if cmp == "a-better" else ("worse" if cmp == "b-better" else "")
        cb = "better" if cmp == "b-better" else ("worse" if cmp == "a-better" else "")
        ind = '<span style="color:#16a34a;font-size:10px">&#9650;</span>' if cmp == "b-better" else ('<span style="color:#FBBF24;font-size:10px">&#9660;</span>' if cmp == "a-better" else "")
        rows += f'<tr><td style="font-weight:600">{b["row"]}</td><td class="{ca}">{va}</td><td class="{cb}">{vb}</td><td style="text-align:center">{ind}</td></tr>'

    html = f"""<!DOCTYPE html><html><head><meta charset="utf-8"><meta name="viewport" content="width=device-width,initial-scale=1">
<title>{plan_a} vs {plan_b}</title><style>
@import url('https://fonts.googleapis.com/css2?family=Inter:wght@300;400;500;600;700;800;900&display=swap');
*{{margin:0;padding:0;box-sizing:border-box}}body{{font-family:'Inter',sans-serif;background:#FAF7F2;color:#262626;font-size:13px;line-height:1.6}}
.c{{max-width:1000px;margin:0 auto;padding:40px 20px}}
.hd{{background:#262626;color:#fff;padding:32px;border-radius:16px;margin-bottom:24px;display:flex;align-items:center;gap:20px}}
.hd img{{height:50px}}.hd h1{{font-weight:800;font-size:24px}}.hd .s{{font-size:12px;opacity:.6;margin-top:4px}}
.sec{{background:#fff;border-radius:14px;padding:24px;margin-bottom:20px;box-shadow:0 1px 3px rgba(0,0,0,.04)}}
table{{width:100%;border-collapse:collapse}}thead th{{background:#262626;color:#fff;padding:10px 12px;text-align:left;font-size:10px;font-weight:700;text-transform:uppercase}}
tbody td{{padding:10px 12px;border-bottom:1px solid #F4F4F6}}tbody tr:hover{{background:#F4F4F6}}
.better{{color:#16a34a;font-weight:700}}.worse{{color:#6B7280}}
.dl{{display:inline-flex;align-items:center;gap:6px;padding:10px 20px;background:#16a34a;color:#fff;border:none;border-radius:8px;font-size:12px;font-weight:700;cursor:pointer;margin-left:auto;text-decoration:none}}
.ft{{text-align:center;color:#6B7280;font-size:10px;margin-top:20px;padding:16px}}
@media print{{.dl{{display:none}}}}
</style></head><body><div class="c">
<div class="hd">{'<img src="data:image/jpeg;base64,'+logo+'">' if logo else ''}
<div><h1>{plan_a} vs {plan_b}</h1><div class="s">Leadway Health 2026 Benefit Comparison | {pd.Timestamp.now().strftime('%d %B %Y')}</div></div>
<button class="dl" onclick="(function(){{var a=document.createElement('a');a.href='data:text/html;charset=utf-8,'+encodeURIComponent(document.documentElement.outerHTML);a.download='Benefit_{plan_a}_vs_{plan_b}.html';a.click();}})()">&#11015; Download</button></div>
<div class="sec"><table><thead><tr><th>Benefit</th><th>{plan_a}</th><th>{plan_b}</th><th></th></tr></thead><tbody>{rows}</tbody></table></div>
<div class="ft">Leadway Health Analytics — Benefit Benchmarking</div></div></body></html>"""

    rp = REPORTS_DIR / f"Benefit_{plan_a}_vs_{plan_b}_{sid}.html"
    rp.write_text(html)
    return redirect(url_for("view_report", filename=rp.name))


@app.route("/benefit-benchmarking/multi", methods=["POST"])
@login_required
def benefit_multi():
    import pandas as pd
    plans = request.form.getlist("plans")
    if len(plans) < 2:
        flash("Please select at least 2 plans.")
        return redirect(url_for("benefit_benchmarking"))

    logo = get_logo_b64()
    sid = str(uuid.uuid4())[:8]
    REPORTS_DIR.mkdir(parents=True, exist_ok=True)

    # Header row
    plan_headers = "".join(f"<th>{p}</th>" for p in plans)

    # Build rows — highlight best value per row
    rows = ""
    for b in LEADWAY_BENEFITS:
        vals = [b.get(p, "—") for p in plans]
        # Find best value
        best_idx = set()
        for i in range(len(vals)):
            is_best = True
            for j in range(len(vals)):
                if i == j: continue
                cmp = _compare_values(vals[i], vals[j])
                if cmp == "b-better":
                    is_best = False
                    break
            if is_best:
                # Verify it's actually better than at least one other
                for j in range(len(vals)):
                    if i == j: continue
                    if _compare_values(vals[j], vals[i]) == "b-better":
                        best_idx.add(i)
                        break

        cells = ""
        for i, v in enumerate(vals):
            cls = ' class="best"' if i in best_idx else ""
            cells += f"<td{cls}>{v}</td>"
        rows += f'<tr><td class="bn">{b["row"]}</td>{cells}</tr>'

    title = " vs ".join(plans)
    html = f"""<!DOCTYPE html><html><head><meta charset="utf-8"><meta name="viewport" content="width=device-width,initial-scale=1">
<title>{title} — Multi-Plan Comparison</title><style>
@import url('https://fonts.googleapis.com/css2?family=Inter:wght@300;400;500;600;700;800;900&display=swap');
*{{margin:0;padding:0;box-sizing:border-box}}body{{font-family:'Inter',sans-serif;background:#FAF7F2;color:#262626;font-size:12px;line-height:1.6}}
.c{{max-width:1200px;margin:0 auto;padding:40px 20px}}
.hd{{background:#262626;color:#fff;padding:32px;border-radius:16px;margin-bottom:24px;display:flex;align-items:center;gap:20px}}
.hd img{{height:50px}}.hd h1{{font-weight:800;font-size:22px}}.hd .s{{font-size:12px;opacity:.6;margin-top:4px}}
.sec{{background:#fff;border-radius:14px;padding:24px;margin-bottom:20px;box-shadow:0 1px 3px rgba(0,0,0,.04);overflow-x:auto}}
table{{width:100%;border-collapse:collapse;min-width:600px}}
thead th{{background:#262626;color:#fff;padding:10px 12px;text-align:left;font-size:10px;font-weight:700;text-transform:uppercase;position:sticky;top:0}}
tbody td{{padding:9px 12px;border-bottom:1px solid #F4F4F6}}tbody tr:hover{{background:#F4F4F6}}
.bn{{font-weight:600;min-width:180px;background:#FAFAFA;position:sticky;left:0}}
.best{{color:#16a34a;font-weight:700;background:#F0FDF4}}
.dl{{display:inline-flex;align-items:center;gap:6px;padding:10px 20px;background:#16a34a;color:#fff;border:none;border-radius:8px;font-size:12px;font-weight:700;cursor:pointer;margin-left:auto;text-decoration:none}}
.ft{{text-align:center;color:#6B7280;font-size:10px;margin-top:20px;padding:16px}}
@media print{{.dl{{display:none}}}}
</style></head><body><div class="c">
<div class="hd">{'<img src="data:image/jpeg;base64,'+logo+'">' if logo else ''}
<div><h1>{title}</h1><div class="s">Leadway Health 2026 Multi-Plan Comparison | {pd.Timestamp.now().strftime('%d %B %Y')}</div></div>
<button class="dl" onclick="(function(){{var a=document.createElement('a');a.href='data:text/html;charset=utf-8,'+encodeURIComponent(document.documentElement.outerHTML);a.download='MultiPlan_Comparison.html';a.click();}})()">&#11015; Download</button></div>
<div class="sec"><table><thead><tr><th>Benefit</th>{plan_headers}</tr></thead><tbody>{rows}</tbody></table></div>
<div class="ft">Leadway Health Analytics — Benefit Benchmarking</div></div></body></html>"""

    rp = REPORTS_DIR / f"Benefit_Multi_{sid}.html"
    rp.write_text(html)
    return redirect(url_for("view_report", filename=rp.name))


@app.route("/benefit-benchmarking/client", methods=["POST"])
@login_required
def benefit_client():
    import pandas as pd
    import re

    client_file = request.files.get("client_benefit")
    if not client_file or client_file.filename == "":
        flash("Please upload the client's benefit schedule.")
        return redirect(url_for("benefit_benchmarking"))

    lw_plan = request.form.get("leadway_plan", "PRO")
    client_name = request.form.get("client_name", "Client").strip() or "Client"
    sid = str(uuid.uuid4())[:8]
    REPORTS_DIR.mkdir(parents=True, exist_ok=True)

    upload_dir = app.config["UPLOAD_FOLDER"] / sid
    upload_dir.mkdir(parents=True, exist_ok=True)
    fpath = upload_dir / client_file.filename
    client_file.save(fpath)

    cdf = pd.read_csv(fpath) if str(fpath).endswith(".csv") else pd.read_excel(fpath)
    cdf.columns = cdf.columns.str.strip()

    # Build client benefit dict
    client_map = {}
    c0, c1 = cdf.columns[0], cdf.columns[1] if len(cdf.columns) > 1 else cdf.columns[0]
    for _, r in cdf.iterrows():
        k = str(r[c0]).strip()
        v = str(r[c1]).strip() if c1 != c0 else "Covered"
        if k and k.lower() != "nan":
            client_map[k.lower()] = {"name": k, "value": v}

    logo = get_logo_b64()
    rows = ""
    lw_wins, cl_wins, same = 0, 0, 0

    for b in LEADWAY_BENEFITS:
        lw_val = b.get(lw_plan, "—")
        cat = b["category"]
        rname = b["row"].lower()

        # Fuzzy match
        cv = "—"
        for ck, cd in client_map.items():
            if cat in ck or ck in cat or any(w in ck for w in rname.split() if len(w) > 3):
                cv = cd["value"]
                break

        cmp = _compare_values(cv, lw_val)
        if cmp == "b-better": lw_wins += 1
        elif cmp == "a-better": cl_wins += 1
        else: same += 1

        ccl = "better" if cmp == "a-better" else ("worse" if cmp == "b-better" else "")
        clw = "better" if cmp == "b-better" else ("worse" if cmp == "a-better" else "")
        badge = ""
        if cmp == "b-better": badge = '<span style="font-size:9px;font-weight:700;padding:2px 6px;border-radius:4px;background:#F0FDF4;color:#16a34a">LEADWAY BETTER</span>'
        elif cmp == "a-better": badge = '<span style="font-size:9px;font-weight:700;padding:2px 6px;border-radius:4px;background:#FFF7ED;color:#F15A24">CLIENT BETTER</span>'

        rows += f'<tr><td style="font-weight:600">{b["row"]}</td><td class="{ccl}">{cv}</td><td class="{clw}">{lw_val}</td><td>{badge}</td></tr>'

    selling = f'<div style="background:#F0FDF4;border:2px solid #16a34a;border-radius:12px;padding:20px;margin-bottom:20px"><div style="color:#16a34a;font-weight:800;font-size:14px;margin-bottom:6px">&#10003; Leadway Advantage</div><div style="font-size:12px;color:#6B7280;line-height:1.7">Leadway {lw_plan} offers <strong>superior coverage on {lw_wins} benefit categories</strong> vs the client\'s current plan.</div></div>' if lw_wins > 0 else ""

    # ── Auto-generate proposal ──
    # Collect where client has better coverage (Leadway gaps to address)
    gaps = []
    strengths = []
    for b in LEADWAY_BENEFITS:
        lw_v = b.get(lw_plan, "—")
        cat = b["category"]
        rname = b["row"]
        cv = "—"
        for ck, cd in client_map.items():
            if cat in ck or ck in cat or any(w in ck for w in rname.lower().split() if len(w) > 3):
                cv = cd["value"]
                break
        cmp = _compare_values(cv, lw_v)
        if cmp == "a-better":
            gaps.append({"benefit": rname, "client": cv, "leadway": lw_v})
        elif cmp == "b-better":
            strengths.append({"benefit": rname, "client": cv, "leadway": lw_v})

    # Find the next tier up that covers more gaps
    plan_idx = ["PLUS", "PRO", "MAX", "PROMAX", "MAGNUM", "MAGNUM PLUS"]
    current_idx = plan_idx.index(lw_plan) if lw_plan in plan_idx else 1
    recommended_plan = lw_plan
    recommended_upgrades = []

    if gaps and current_idx < len(plan_idx) - 1:
        # Check if upgrading to next tier covers gaps
        next_plan = plan_idx[current_idx + 1]
        gaps_fixed = 0
        for g in gaps:
            for b in LEADWAY_BENEFITS:
                if b["row"] == g["benefit"]:
                    next_val = b.get(next_plan, "—")
                    cmp_next = _compare_values(g["client"], next_val)
                    if cmp_next != "a-better":
                        gaps_fixed += 1
                    break
        if gaps_fixed > len(gaps) * 0.5:
            recommended_plan = next_plan
        else:
            recommended_plan = lw_plan

    # Build proposal HTML
    proposal_items = ""
    if strengths:
        items = "".join(f"<li><strong>{s['benefit']}</strong>: Leadway offers {s['leadway']} vs client's {s['client']}</li>" for s in strengths[:5])
        proposal_items += f'<div style="margin-bottom:16px"><div style="font-weight:700;font-size:13px;color:#16a34a;margin-bottom:6px">Key Selling Points</div><ul style="font-size:12px;color:#4B5563;padding-left:20px;line-height:1.8">{items}</ul></div>'

    if gaps:
        items = "".join(f"<li><strong>{g['benefit']}</strong>: Client currently has {g['client']}, Leadway {lw_plan} offers {g['leadway']}. Consider upgrading this benefit or moving to {recommended_plan}.</li>" for g in gaps[:5])
        proposal_items += f'<div style="margin-bottom:16px"><div style="font-weight:700;font-size:13px;color:#F15A24;margin-bottom:6px">Coverage Gaps to Address</div><ul style="font-size:12px;color:#4B5563;padding-left:20px;line-height:1.8">{items}</ul></div>'

    rec_premium = ""
    for b in LEADWAY_BENEFITS:
        if b["category"] == "premium":
            rec_premium = b.get(recommended_plan, "—")
            break

    proposal_recommendation = f"""<div style="background:#EFF6FF;border:2px solid #1B1464;border-radius:12px;padding:20px;margin-top:16px">
    <div style="font-weight:800;font-size:14px;color:#1B1464;margin-bottom:8px">&#128203; Recommendation</div>
    <div style="font-size:13px;color:#4B5563;line-height:1.8">
    Based on this analysis, we recommend <strong>Leadway {recommended_plan}</strong> for {client_name}.
    {"This plan covers all the client's current benefits and offers superior coverage on " + str(lw_wins) + " additional categories." if recommended_plan == lw_plan else "Upgrading from " + lw_plan + " to <strong>" + recommended_plan + "</strong> would address " + str(len(gaps)) + " coverage gaps where the client currently has better benefits."}
    <br><br>
    <strong>Recommended Premium:</strong> &#8358;{rec_premium} per individual per annum.
    </div></div>"""

    proposal_html = f"""<div style="background:#fff;border-radius:14px;padding:24px;margin-bottom:20px;box-shadow:0 1px 3px rgba(0,0,0,.04)">
    <div style="font-weight:700;font-size:18px;color:#262626;margin-bottom:16px;padding-bottom:8px;border-bottom:2px solid #E6E6E6">&#128220; Auto-Generated <span style="color:#16a34a">Proposal</span></div>
    <div style="font-size:12px;color:#6B7280;margin-bottom:16px">Based on the benefit comparison for {client_name}</div>
    {proposal_items}
    {proposal_recommendation}
    </div>"""

    html = f"""<!DOCTYPE html><html><head><meta charset="utf-8"><meta name="viewport" content="width=device-width,initial-scale=1">
<title>{client_name} vs Leadway {lw_plan}</title><style>
@import url('https://fonts.googleapis.com/css2?family=Inter:wght@300;400;500;600;700;800;900&display=swap');
*{{margin:0;padding:0;box-sizing:border-box}}body{{font-family:'Inter',sans-serif;background:#FAF7F2;color:#262626;font-size:13px;line-height:1.6}}
.c{{max-width:1100px;margin:0 auto;padding:40px 20px}}
.hd{{background:#262626;color:#fff;padding:32px;border-radius:16px;margin-bottom:24px;display:flex;align-items:center;gap:20px}}
.hd img{{height:50px}}.hd h1{{font-weight:800;font-size:22px}}.hd .s{{font-size:12px;opacity:.6;margin-top:4px}}
.kg{{display:grid;grid-template-columns:repeat(3,1fr);gap:14px;margin-bottom:24px}}
.kp{{background:#fff;border-radius:12px;padding:20px;text-align:center;box-shadow:0 1px 3px rgba(0,0,0,.04)}}
.kp .lb{{font-size:10px;font-weight:700;text-transform:uppercase;color:#6B7280;margin-bottom:4px}}
.kp .vl{{font-weight:800;font-size:24px}}
.kp.g .vl{{color:#16a34a}}.kp.o .vl{{color:#F15A24}}
.sec{{background:#fff;border-radius:14px;padding:24px;margin-bottom:20px;box-shadow:0 1px 3px rgba(0,0,0,.04)}}
table{{width:100%;border-collapse:collapse}}thead th{{background:#262626;color:#fff;padding:10px 12px;text-align:left;font-size:10px;font-weight:700;text-transform:uppercase}}
tbody td{{padding:10px 12px;border-bottom:1px solid #F4F4F6}}tbody tr:hover{{background:#F4F4F6}}
.better{{color:#16a34a;font-weight:700}}.worse{{color:#6B7280}}
.dl{{display:inline-flex;align-items:center;gap:6px;padding:10px 20px;background:#16a34a;color:#fff;border:none;border-radius:8px;font-size:12px;font-weight:700;cursor:pointer;margin-left:auto;text-decoration:none}}
.ft{{text-align:center;color:#6B7280;font-size:10px;margin-top:20px;padding:16px}}
@media print{{.dl{{display:none}}}}
</style></head><body><div class="c">
<div class="hd">{'<img src="data:image/jpeg;base64,'+logo+'">' if logo else ''}
<div><h1>{client_name} vs Leadway {lw_plan}</h1><div class="s">Benefit Benchmarking & Proposal | {pd.Timestamp.now().strftime('%d %B %Y')}</div></div>
<button class="dl" onclick="(function(){{var a=document.createElement('a');a.href='data:text/html;charset=utf-8,'+encodeURIComponent(document.documentElement.outerHTML);a.download='Benefit_{client_name.replace(' ','_')}_vs_{lw_plan}.html';a.click();}})()">&#11015; Download</button></div>
<div class="kg"><div class="kp g"><div class="lb">Leadway Offers More</div><div class="vl">{lw_wins}</div></div><div class="kp o"><div class="lb">Client Has More</div><div class="vl">{cl_wins}</div></div><div class="kp"><div class="lb">Equivalent</div><div class="vl">{same}</div></div></div>
{selling}
<div class="sec"><table><thead><tr><th>Benefit</th><th>{client_name} (Current)</th><th>Leadway {lw_plan}</th><th></th></tr></thead><tbody>{rows}</tbody></table></div>
{proposal_html}
<div class="ft">Leadway Health Analytics — Benefit Benchmarking & Proposal</div></div></body></html>"""

    cn = client_name.replace(" ", "_")[:20]
    rp = REPORTS_DIR / f"Benefit_{cn}_vs_{lw_plan}_{sid}.html"
    rp.write_text(html)
    return redirect(url_for("view_report", filename=rp.name))


@app.route("/benefit-benchmarking/upload", methods=["POST"])
@login_required
def benefit_benchmarking_upload():
    return redirect(url_for("benefit_benchmarking"))


# ═══════════════════════════════════════════════
# MODULE 4 — Pricing Tool
# ═══════════════════════════════════════════════

@app.route("/pricing-tool")
@login_required
def pricing_tool():
    logo = get_logo_b64()
    report_list = list_reports("Pricing_")
    return render_template("pricing_tool.html", logo=logo, reports=report_list)


@app.route("/pricing-tool/upload", methods=["POST"])
@login_required
def pricing_tool_upload():
    return redirect(url_for("pricing_tool_calculate"))


@app.route("/pricing-tool/calculate", methods=["POST"])
@login_required
def pricing_tool_calculate():
    logo = get_logo_b64()

    # ── Plan base premiums (from 2026 Corporate benefit table) ──
    PLAN_DATA = {
        "PLUS":        {"individual": 121268,  "family": 576023,    "tier": "D"},
        "PRO":         {"individual": 164200,  "family": 779950,    "tier": "C+D"},
        "MAX":         {"individual": 262275,  "family": 1245806,   "tier": "B+C+D"},
        "PROMAX":      {"individual": 382582,  "family": 1817265,   "tier": "A+B+C+D"},
        "MAGNUM":      {"individual": 734250,  "family": 3487688,   "tier": "A+B+C+D"},
        "MAGNUM PLUS": {"individual": 1737750, "family": 8254313,   "tier": "A*+A+B+C+D"},
    }

    # ── Industry risk loadings ──
    INDUSTRY_LOADING = {
        "oil_gas": 1.20, "manufacturing": 1.15, "government": 1.10, "banking": 1.05,
        "telecom": 1.05, "fmcg": 1.05, "education": 1.00, "tech": 1.00,
        "consulting": 1.00, "ngo": 1.00, "other": 1.00,
    }

    # ── Benefit upgrade cost factors (actuarial estimates per member) ──
    # Cost = (new_limit - base_limit) * utilization_rate * avg_cost_factor
    BENEFIT_DEFAULTS = {
        "dental":        {"base": 50000,  "util": 0.35, "cost_factor": 0.08},
        "optical":       {"base": 30000,  "util": 0.25, "cost_factor": 0.06},
        "surgery":       {"base": 250000, "util": 0.08, "cost_factor": 0.12},
        "icu_days":      {"base": 3,      "util": 0.03, "per_day_cost": 85000},
        "major_disease": {"base": 500000, "util": 0.02, "cost_factor": 0.15},
    }
    NICU_COST = 45000       # per member loading for NICU
    GYM_COST = 8000         # per member
    IMMUNIZATION_COST = 5000  # per member

    # ── Read form ──
    company_name = request.form.get("company_name", "Client")
    industry = request.form.get("industry", "other")
    plan_name = request.form.get("plan", "PRO")
    total_principals = int(request.form.get("total_principals", 50))
    total_dependants = int(request.form.get("total_dependants", 0))
    total_lives = total_principals + total_dependants

    plan = PLAN_DATA.get(plan_name, PLAN_DATA["PRO"])
    base_price = plan["individual"]

    # ── Calculate benefit upgrades ──
    upgrade_total = 0
    comparison = []
    upgrade_count = 0

    # Dental
    dental_on = request.form.get("dental_toggle") == "on"
    dental_limit = int(request.form.get("dental_limit", 50000))
    d = BENEFIT_DEFAULTS["dental"]
    dental_cost = 0
    if dental_on and dental_limit > d["base"]:
        dental_cost = (dental_limit - d["base"]) * d["util"] * d["cost_factor"]
        upgrade_total += dental_cost
        upgrade_count += 1
    comparison.append({"name": "Dental", "base": f"\u20A6{d['base']:,.0f}", "custom": f"\u20A6{dental_limit:,.0f}" if dental_on else f"\u20A6{d['base']:,.0f}", "impact": f"+\u20A6{dental_cost:,.0f}" if dental_cost > 0 else "\u2014"})

    # Optical
    optical_on = request.form.get("optical_toggle") == "on"
    optical_limit = int(request.form.get("optical_limit", 30000))
    d = BENEFIT_DEFAULTS["optical"]
    optical_cost = 0
    if optical_on and optical_limit > d["base"]:
        optical_cost = (optical_limit - d["base"]) * d["util"] * d["cost_factor"]
        upgrade_total += optical_cost
        upgrade_count += 1
    comparison.append({"name": "Optical", "base": f"\u20A6{d['base']:,.0f}", "custom": f"\u20A6{optical_limit:,.0f}" if optical_on else f"\u20A6{d['base']:,.0f}", "impact": f"+\u20A6{optical_cost:,.0f}" if optical_cost > 0 else "\u2014"})

    # Surgery
    surgery_on = request.form.get("surgery_toggle") == "on"
    surgery_limit = int(request.form.get("surgery_limit", 250000))
    d = BENEFIT_DEFAULTS["surgery"]
    surgery_cost = 0
    if surgery_on and surgery_limit > d["base"]:
        surgery_cost = (surgery_limit - d["base"]) * d["util"] * d["cost_factor"]
        upgrade_total += surgery_cost
        upgrade_count += 1
    comparison.append({"name": "Surgery", "base": f"\u20A6{d['base']:,.0f}", "custom": f"\u20A6{surgery_limit:,.0f}" if surgery_on else f"\u20A6{d['base']:,.0f}", "impact": f"+\u20A6{surgery_cost:,.0f}" if surgery_cost > 0 else "\u2014"})

    # ICU
    icu_on = request.form.get("icu_toggle") == "on"
    icu_days = int(request.form.get("icu_days", 3))
    d = BENEFIT_DEFAULTS["icu_days"]
    icu_cost = 0
    if icu_on and icu_days > d["base"]:
        icu_cost = (icu_days - d["base"]) * d["util"] * d["per_day_cost"]
        upgrade_total += icu_cost
        upgrade_count += 1
    comparison.append({"name": "ICU Days", "base": f"{d['base']} days", "custom": f"{icu_days} days" if icu_on else f"{d['base']} days", "impact": f"+\u20A6{icu_cost:,.0f}" if icu_cost > 0 else "\u2014"})

    # NICU
    nicu_on = request.form.get("nicu_toggle") == "on"
    nicu_cost = NICU_COST if nicu_on else 0
    upgrade_total += nicu_cost
    if nicu_on:
        upgrade_count += 1
    comparison.append({"name": "NICU", "base": "Not included", "custom": "Included" if nicu_on else "Not included", "impact": f"+\u20A6{nicu_cost:,.0f}" if nicu_cost > 0 else "\u2014"})

    # Major Disease
    major_on = request.form.get("major_disease_toggle") == "on"
    major_limit = int(request.form.get("major_disease_limit", 500000))
    d = BENEFIT_DEFAULTS["major_disease"]
    major_cost = 0
    if major_on and major_limit > d["base"]:
        major_cost = (major_limit - d["base"]) * d["util"] * d["cost_factor"]
        upgrade_total += major_cost
        upgrade_count += 1
    comparison.append({"name": "Major Disease", "base": f"\u20A6{d['base']:,.0f}", "custom": f"\u20A6{major_limit:,.0f}" if major_on else f"\u20A6{d['base']:,.0f}", "impact": f"+\u20A6{major_cost:,.0f}" if major_cost > 0 else "\u2014"})

    # Gym
    gym_on = request.form.get("gym_toggle") == "on"
    gym_cost = GYM_COST if gym_on else 0
    upgrade_total += gym_cost
    comparison.append({"name": "Gym / Wellness", "base": "Not included", "custom": "Included" if gym_on else "Not included", "impact": f"+\u20A6{gym_cost:,.0f}" if gym_cost > 0 else "\u2014"})

    # Immunization
    immun_on = request.form.get("immunization_toggle") == "on"
    immun_cost = IMMUNIZATION_COST if immun_on else 0
    upgrade_total += immun_cost
    comparison.append({"name": "Immunization", "base": "Not included", "custom": "Included" if immun_on else "Not included", "impact": f"+\u20A6{immun_cost:,.0f}" if immun_cost > 0 else "\u2014"})

    # ── Industry loading ──
    ind_multiplier = INDUSTRY_LOADING.get(industry, 1.0)
    subtotal_before_loading = base_price + upgrade_total
    industry_loading = subtotal_before_loading * (ind_multiplier - 1)

    # ── Small group loading (< 50 lives = +10%, < 30 lives = +15%) ──
    if total_lives < 30:
        small_group_pct = 0.15
    elif total_lives < 50:
        small_group_pct = 0.10
    else:
        small_group_pct = 0.0
    small_group_loading = subtotal_before_loading * small_group_pct

    # ── Multi-upgrade loading (3+ upgrades = +3%, 5+ = +5%) ──
    if upgrade_count >= 5:
        multi_pct = 0.05
    elif upgrade_count >= 3:
        multi_pct = 0.03
    else:
        multi_pct = 0.0
    multi_loading = subtotal_before_loading * multi_pct
    upgrade_total += multi_loading  # fold into upgrades display

    # ── Volume discount (100+ lives = 3%, 200+ = 5%, 500+ = 7%) ──
    if total_lives >= 500:
        discount_pct = 0.07
    elif total_lives >= 200:
        discount_pct = 0.05
    elif total_lives >= 100:
        discount_pct = 0.03
    else:
        discount_pct = 0.0

    gross_price = base_price + upgrade_total + industry_loading + small_group_loading
    discount = gross_price * discount_pct
    final_price = gross_price - discount
    family_price = final_price * (plan["family"] / plan["individual"])
    annual_total = (final_price * total_principals) + (family_price * total_dependants) if total_dependants > 0 else final_price * total_lives

    # ── Risk assessment ──
    high_risk = (surgery_on and icu_on and plan_name in ("PROMAX", "MAGNUM", "MAGNUM PLUS")) or nicu_on
    moderate_risk = upgrade_count >= 3 or (surgery_on and surgery_limit >= 500000)
    if high_risk:
        risk_level, risk_label = "high", "High Risk — Manual review recommended"
    elif moderate_risk:
        risk_level, risk_label = "moderate", "Moderate Risk — Monitor at renewal"
    else:
        risk_level, risk_label = "low", "Low Risk — Standard terms"

    # ── Underwriting decision ──
    if high_risk:
        decision = "Escalate"
        decision_class = "escalate"
        uw_notes = f"This quote for <strong>{company_name}</strong> includes high-cost benefits "
        reasons = []
        if nicu_on:
            reasons.append("NICU coverage")
        if surgery_on and icu_on and plan_name in ("PROMAX", "MAGNUM", "MAGNUM PLUS"):
            reasons.append(f"Surgery (\u20A6{surgery_limit:,.0f}) + ICU ({icu_days} days) on {plan_name}")
        uw_notes += "(" + ", ".join(reasons) + "). "
        uw_notes += "Recommend pricing review, benefit cap, or co-payment arrangement before binding."
    elif moderate_risk:
        decision = "Review Required"
        decision_class = "review"
        uw_notes = f"<strong>{company_name}</strong> has {upgrade_count} benefit upgrades. "
        if surgery_on and surgery_limit >= 500000:
            uw_notes += f"Surgery limit increased to \u20A6{surgery_limit:,.0f} (high cost driver). "
        uw_notes += "Review claims history at renewal. Consider adding waiting periods for upgraded benefits."
    else:
        decision = "Approved"
        decision_class = "approve"
        uw_notes = f"Standard plan configuration for <strong>{company_name}</strong>. No elevated risk factors detected. Quote can be issued at standard terms."

    # ── Format helpers ──
    def fmt(v):
        return f"\u20A6{v:,.0f}"

    ind_label = {
        "oil_gas": "Oil & Gas +20%", "manufacturing": "Manufacturing +15%",
        "government": "Government +10%", "banking": "Banking +5%",
        "telecom": "Telecom +5%", "fmcg": "FMCG +5%",
    }.get(industry, "Standard 0%")

    result = {
        "company_name": company_name,
        "plan_name": plan_name,
        "total_lives": total_lives,
        "provider_tier": plan["tier"],
        "comparison": comparison,
        "base_price_fmt": fmt(base_price),
        "upgrades_fmt": fmt(upgrade_total),
        "industry_loading_pct": ind_label,
        "industry_loading_fmt": fmt(industry_loading),
        "small_group_fmt": fmt(small_group_loading),
        "discount": discount,
        "discount_fmt": fmt(discount),
        "final_price_fmt": fmt(final_price),
        "family_price_fmt": fmt(family_price),
        "annual_total_fmt": fmt(annual_total),
        "risk_level": risk_level,
        "risk_label": risk_label,
        "decision": decision,
        "decision_class": decision_class,
        "uw_notes": uw_notes,
    }

    return render_template("pricing_tool.html", logo=logo, result=result, reports=list_reports("Pricing_"))


# ═══════════════════════════════════════════════
# ENROLMENT TEMPLATE GENERATOR
# ═══════════════════════════════════════════════

@app.route("/enrolment-generator")
@login_required
def enrolment_generator():
    logo = get_logo_b64()
    report_list = list_reports("Enrolment_", ext="xlsx")
    return render_template("enrolment_generator.html", logo=logo, reports=report_list)


@app.route("/enrolment-generator/upload", methods=["POST"])
@login_required
def enrolment_generator_upload():
    import pandas as pd
    import numpy as np
    from datetime import datetime
    import io

    hr_file = request.files.get("hr_file")
    if not hr_file or hr_file.filename == "":
        flash("Please upload an HR / enrolment file.")
        return redirect(url_for("enrolment_generator"))

    client_name = request.form.get("client_name", "Client").strip() or "Client"
    session_id = str(uuid.uuid4())[:8]
    upload_dir = app.config["UPLOAD_FOLDER"] / session_id
    upload_dir.mkdir(parents=True, exist_ok=True)
    fpath = upload_dir / hr_file.filename
    hr_file.save(fpath)

    # Parse file
    if str(fpath).endswith(".csv"):
        df = pd.read_csv(fpath)
    else:
        df = pd.read_excel(fpath)
    df.columns = [str(c).strip() for c in df.columns]

    # ── STEP 0: Detect format and expand ──
    # Clean column names: strip whitespace, non-breaking spaces, newlines
    df.columns = [str(c).replace("\xa0", " ").replace("\n", " ").replace("\r", " ").strip() for c in df.columns]
    cols = list(df.columns)
    cols_lower = [c.lower().strip() for c in cols]

    # Detect wide format: look for repeated First Name / Last Name patterns
    fn_cols = [i for i, c in enumerate(cols_lower) if "first name" in c]
    ln_cols = [i for i, c in enumerate(cols_lower) if "last name" in c]
    is_wide = len(fn_cols) > 1 or len(ln_cols) > 1

    # Also detect: Spouse/Child/Dep columns
    if not is_wide:
        is_wide = any(any(kw in c for kw in ["spouse", "child1", "child 1", "dep1", "dep 1"]) for c in cols_lower)

    expanded_rows = []

    def clean_v(v):
        """Clean a cell value."""
        if v is None: return ""
        s = str(v).strip()
        if s.lower() in ("nan", "none", "nat", "", "nil", "n/a", "na", "-", "null"): return ""
        return s

    def fmt_phone(v):
        """Format phone number."""
        s = clean_v(v)
        if not s: return ""
        # Remove country code prefix, keep as string
        s = s.replace("+", "").replace(" ", "").replace("-", "")
        try:
            n = int(float(s))
            s = str(n)
        except (ValueError, TypeError):
            return s
        # Nigerian numbers: if starts with 234, strip to 0...
        if s.startswith("234") and len(s) > 10:
            s = "0" + s[3:]
        elif not s.startswith("0") and len(s) == 10:
            s = "0" + s
        return s

    def find_col(keywords, start_idx=0, end_idx=None):
        """Find a column index by keywords within a range."""
        end = end_idx or len(cols)
        for i in range(start_idx, end):
            cl = cols_lower[i]
            for kw in keywords:
                if kw in cl:
                    return i
        return None

    if is_wide:
        # ── WIDE FORMAT: Group columns into blocks ──
        # Strategy: find all "first name" columns — each one starts a person block
        # The principal block is the first one, subsequent blocks are dependants

        # Find principal columns (before the first repeated first name)
        # Principal might use "First Name" (col 3) and "Last Name" (col 4)
        p_fn = fn_cols[0] if fn_cols else None
        p_ln = ln_cols[0] if ln_cols else None

        # Find other principal fields by scanning cols before the first dependant block
        first_dep_start = fn_cols[1] if len(fn_cols) > 1 else (ln_cols[1] if len(ln_cols) > 1 else len(cols))
        p_email = find_col(["email"], 0, first_dep_start)
        p_phone = find_col(["phone", "contact", "mobile", "tel"], 0, first_dep_start)
        p_dob = find_col(["date of birth", "dob", "birth"], 0, first_dep_start)
        p_empid = find_col(["employee id", "staff id", "emp id"], 0, first_dep_start)
        p_addr = find_col(["address"], 0, first_dep_start)
        p_name = find_col(["name"], 0, first_dep_start)  # Full name column

        # Build dependant blocks by pairing each remaining First Name col
        # with its nearest Last Name col, then scanning for DOB/Phone/EmpID nearby
        dep_blocks = []
        used_indices = set()
        if p_fn is not None: used_indices.add(p_fn)
        if p_ln is not None: used_indices.add(p_ln)
        remaining_fn = [i for i in fn_cols if i not in used_indices]

        for fi in remaining_fn:
            # Find nearest Last Name col (prefer fi+1, then fi-1, then next available)
            paired_ln = None
            for li in ln_cols:
                if li == fi + 1:
                    paired_ln = li; break
                elif li == fi - 1:
                    paired_ln = li; break
            if paired_ln is None:
                for li in sorted(ln_cols):
                    if li > fi and li not in used_indices:
                        paired_ln = li; break

            # Block range: from min(fn,ln) to next fn_col or end
            block_start = min(fi, paired_ln) if paired_ln else fi
            next_fn = None
            for nf in remaining_fn:
                if nf > fi:
                    next_fn = nf; break
            block_end = next_fn if next_fn else len(cols)

            block = {"fn": fi, "ln": paired_ln, "dob": None, "phone": None, "empid": None, "addr": None}
            if paired_ln is not None:
                used_indices.add(paired_ln)

            # Scan the block range for DOB/Phone/EmpID
            for j in range(block_start, block_end):
                if j == fi or j == paired_ln:
                    continue
                cl = cols_lower[j]
                if ("date of birth" in cl or "dob" in cl) and block["dob"] is None: block["dob"] = j
                elif any(k in cl for k in ["phone", "contact", "mobile"]) and block["phone"] is None: block["phone"] = j
                elif any(k in cl for k in ["employee id", "staff id"]) and block["empid"] is None: block["empid"] = j
                elif "address" in cl and block["addr"] is None: block["addr"] = j
            dep_blocks.append(block)

        # Now iterate rows and expand
        for _, row in df.iterrows():
            # Principal
            fn = clean_v(row.iloc[p_fn]) if p_fn is not None else ""
            ln = clean_v(row.iloc[p_ln]) if p_ln is not None else ""
            full = clean_v(row.iloc[p_name]) if p_name is not None else ""
            if not fn and not ln and full:
                parts = full.split()
                fn = parts[0] if parts else ""
                ln = parts[-1] if len(parts) > 1 else ""
            if not fn and not ln:
                continue

            email = clean_v(row.iloc[p_email]) if p_email is not None else ""
            phone = fmt_phone(row.iloc[p_phone]) if p_phone is not None else ""
            dob = row.iloc[p_dob] if p_dob is not None else ""
            empid = clean_v(row.iloc[p_empid]) if p_empid is not None else ""
            addr = clean_v(row.iloc[p_addr]) if p_addr is not None else ""

            expanded_rows.append({
                "Firstname": fn.upper(), "Surname": ln.upper(), "OtherName": "",
                "DOB": dob, "Gender": "", "Relationship": "SELF", "Member_Dep": "O",
                "contacts": phone, "email": email, "staff_id": empid,
                "address": addr.replace(",", " "), "city": "",
                "marital": "", "title": "",
                "plan_code": "", "group_code": "", "start_date": "", "state": "",
            })

            # Dependants — deduplicate by (firstname, lastname, dob)
            seen_deps = set()
            for dep in dep_blocks:
                d_fn = clean_v(row.iloc[dep["fn"]]) if dep["fn"] is not None else ""
                d_ln = clean_v(row.iloc[dep["ln"]]) if dep["ln"] is not None else ""
                if not d_fn and not d_ln:
                    continue
                d_dob = row.iloc[dep["dob"]] if dep["dob"] is not None else ""
                d_dob_str = str(d_dob)[:10] if d_dob else ""

                # Deduplicate
                dep_key = (d_fn.upper(), d_ln.upper(), d_dob_str)
                if dep_key in seen_deps:
                    continue
                seen_deps.add(dep_key)

                d_phone = fmt_phone(row.iloc[dep["phone"]]) if dep["phone"] is not None else ""
                d_empid = clean_v(row.iloc[dep["empid"]]) if dep["empid"] is not None else empid

                # Do NOT assume gender or title — leave blank
                expanded_rows.append({
                    "Firstname": d_fn.upper(), "Surname": d_ln.upper(), "OtherName": "",
                    "DOB": d_dob, "Gender": "", "Relationship": "", "Member_Dep": "D",
                    "contacts": d_phone, "email": "", "staff_id": d_empid,
                    "address": addr.replace(",", " "), "city": "LAGOS",
                    "marital": "", "title": "",
                    "plan_code": "", "group_code": "", "start_date": "", "state": "",
                })

    else:
        # ── STANDARD FORMAT ──
        field_map = {}
        for i, cl in enumerate(cols_lower):
            if "surname" in cl or "last name" in cl:
                if "surname" not in field_map: field_map["surname"] = cols[i]
            elif "firstname" in cl or "first name" in cl:
                if "firstname" not in field_map: field_map["firstname"] = cols[i]
            elif "other name" in cl or "middle" in cl:
                if "other_name" not in field_map: field_map["other_name"] = cols[i]
            elif "name" in cl and "surname" not in field_map and "firstname" not in field_map:
                if "name" not in field_map: field_map["name"] = cols[i]
            elif "dob" in cl or "date of birth" in cl or "birth date" in cl:
                if "dob" not in field_map: field_map["dob"] = cols[i]
            elif "gender" in cl or "sex" in cl:
                if "gender" not in field_map: field_map["gender"] = cols[i]
            elif "relationship" in cl or "relation" in cl:
                if "rel" not in field_map: field_map["rel"] = cols[i]
            elif "title" in cl:
                if "title" not in field_map: field_map["title"] = cols[i]
            elif "address" in cl:
                if "address" not in field_map: field_map["address"] = cols[i]
            elif "city" in cl or "town" in cl:
                if "city" not in field_map: field_map["city"] = cols[i]
            elif "state" in cl:
                if "state" not in field_map: field_map["state"] = cols[i]
            elif "staff" in cl or "employee" in cl or "emp" in cl:
                if "staff_id" not in field_map: field_map["staff_id"] = cols[i]
            elif "marital" in cl:
                if "marital" not in field_map: field_map["marital"] = cols[i]
            elif "phone" in cl or "contact" in cl or "mobile" in cl:
                if "contacts" not in field_map: field_map["contacts"] = cols[i]
            elif "plan" in cl:
                if "plan_code" not in field_map: field_map["plan_code"] = cols[i]
            elif "email" in cl:
                if "email" not in field_map: field_map["email"] = cols[i]
            elif "group" in cl:
                if "group_code" not in field_map: field_map["group_code"] = cols[i]
            elif "start date" in cl or "effective" in cl:
                if "start_date" not in field_map: field_map["start_date"] = cols[i]
            elif "member" in cl and ("dep" in cl or "indicator" in cl):
                if "member_dep" not in field_map: field_map["member_dep"] = cols[i]

        for _, row in df.iterrows():
            if "surname" in field_map and "firstname" in field_map:
                sn = clean_v(row.get(field_map["surname"], ""))
                fn = clean_v(row.get(field_map["firstname"], ""))
                on = clean_v(row.get(field_map.get("other_name", ""), ""))
            elif "name" in field_map:
                full = clean_v(row.get(field_map["name"], ""))
                parts = full.upper().split()
                sn = parts[-1] if len(parts) > 1 else (parts[0] if parts else "")
                fn = parts[0] if parts else ""
                on = " ".join(parts[1:-1]) if len(parts) > 2 else ""
            else:
                continue
            if not fn and not sn:
                continue

            rel = clean_v(row.get(field_map.get("rel", ""), "")).upper()
            md = clean_v(row.get(field_map.get("member_dep", ""), "")).upper()
            if not md:
                md = "O" if rel in ("SELF", "PRINCIPAL", "MAIN MEMBER", "") else "D"
            if not rel:
                rel = "SELF" if md == "O" else "DEPENDENT"

            expanded_rows.append({
                "Firstname": fn.upper(), "Surname": sn.upper(), "OtherName": on.upper(),
                "DOB": row.get(field_map.get("dob", ""), ""),
                "Gender": clean_v(row.get(field_map.get("gender", ""), "")).upper(),
                "Relationship": rel, "Member_Dep": md,
                "title": clean_v(row.get(field_map.get("title", ""), "")).upper(),
                "address": clean_v(row.get(field_map.get("address", ""), "")).replace(",", " "),
                "city": clean_v(row.get(field_map.get("city", ""), "")).upper(),
                "state": clean_v(row.get(field_map.get("state", ""), "")).upper(),
                "staff_id": clean_v(row.get(field_map.get("staff_id", ""), "")),
                "marital": clean_v(row.get(field_map.get("marital", ""), "")).upper(),
                "contacts": fmt_phone(row.get(field_map.get("contacts", ""), "")),
                "plan_code": clean_v(row.get(field_map.get("plan_code", ""), "")),
                "email": clean_v(row.get(field_map.get("email", ""), "")),
                "group_code": clean_v(row.get(field_map.get("group_code", ""), "")),
                "start_date": clean_v(row.get(field_map.get("start_date", ""), "")),
            })

    if not expanded_rows:
        flash("No valid records found. Check your file format.")
        return redirect(url_for("enrolment_generator"))

    result = pd.DataFrame(expanded_rows)

    # ── Name cleaning (already split in expansion, just ensure uppercase) ──
    def clean_val(v):
        s = str(v).strip()
        return "" if s.lower() in ("nan", "none", "nat") else s

    result["Surname"] = result["Surname"].apply(lambda x: clean_val(x).upper())
    result["Firstname"] = result["Firstname"].apply(lambda x: clean_val(x).upper())
    result["OtherName"] = result.get("OtherName", pd.Series([""] * len(result))).apply(lambda x: clean_val(x).upper())

    # ── STEP 4: Date standardization ──
    def parse_dob(val):
        v = clean_val(val)
        if not v or v == "############":
            return pd.NaT
        try:
            num = float(v.replace(",", ""))
            if 10000 < num < 60000:
                return pd.to_datetime("1899-12-30") + pd.to_timedelta(num, unit="D")
        except (ValueError, TypeError):
            pass
        try:
            return pd.to_datetime(v, dayfirst=True)
        except (ValueError, TypeError):
            return pd.NaT

    result["DOB_Parsed"] = result["DOB"].apply(parse_dob)
    result["DOB_Formatted"] = result["DOB_Parsed"].apply(lambda d: d.strftime("%d/%m/%Y") if pd.notna(d) else "")

    def parse_start_date(val):
        v = clean_val(val)
        if not v: return ""
        try:
            d = pd.to_datetime(v, dayfirst=True)
            return d.strftime("%d/%m/%Y") if pd.notna(d) else ""
        except (ValueError, TypeError):
            return ""

    result["Start_Formatted"] = result.get("start_date", pd.Series([""] * len(result))).apply(parse_start_date)

    # ── STEP 5: Age calculation + Gender/Title inference ──
    today = datetime.now()
    result["Age"] = result["DOB_Parsed"].apply(lambda d: (today - d).days // 365 if pd.notna(d) else None)

    def infer_title(row):
        t = clean_val(row.get("title", "")).upper()
        if t and t not in ("NAN", ""): return t
        # Only infer from explicit relationship — never assume
        rel = str(row.get("Relationship", "")).upper()
        gender = clean_val(row.get("Gender", "")).upper()
        if rel == "SON": return "MSTR"
        if rel == "DAUGHTER": return "MISS"
        if rel == "SPOUSE" and gender == "FEMALE": return "MRS"
        if rel == "SPOUSE" and gender == "MALE": return "MR"
        if rel == "SELF" and gender == "MALE": return "MR"
        if rel == "SELF" and gender == "FEMALE": return "MRS"
        return ""  # Don't assume

    def infer_gender(row):
        g = clean_val(row.get("Gender", "")).upper()
        if g in ("MALE", "M"): return "MALE"
        if g in ("FEMALE", "F"): return "FEMALE"
        # Only infer from explicit relationship
        rel = str(row.get("Relationship", "")).upper()
        if rel == "DAUGHTER": return "FEMALE"
        if rel == "SON": return "MALE"
        if rel == "SPOUSE": return ""  # Don't assume spouse gender
        return ""

    result["Title"] = result.apply(infer_title, axis=1)
    result["Gender_Clean"] = result.apply(infer_gender, axis=1)

    def infer_marital(row):
        m = clean_val(row.get("marital", "")).upper()
        if m and m not in ("NAN", ""): return m
        rel = str(row.get("Relationship", "")).upper()
        if rel == "SELF": return "MARRIED"
        if rel == "SPOUSE": return "MARRIED"
        if rel in ("SON", "DAUGHTER"): return "SINGLE"
        return ""

    result["Marital_Clean"] = result.apply(infer_marital, axis=1)

    # ── STEP 6: Validation ──
    validation_notes = []
    validation_flags = []
    for _, row in result.iterrows():
        nl = []
        flag = "VALID"
        rel = str(row["Relationship"]).upper()
        age = row["Age"]

        if pd.isna(row["DOB_Parsed"]):
            nl.append("Missing DOB")
            flag = "WARNING"
        if age is not None:
            if rel in ("SELF", "SPOUSE", "") and age > 64:
                nl.append(f"Overaged {rel} (age {age})")
                flag = "RED"
            elif rel in ("SON", "DAUGHTER", "DEPENDENT") and age > 21:
                nl.append(f"Overaged Dependent (age {age})")
                flag = "RED"
        validation_notes.append("; ".join(nl) if nl else "")
        validation_flags.append(flag)

    result["Notes"] = validation_notes
    result["Flag"] = validation_flags

    dup_mask = result.duplicated(subset=["Surname", "Firstname", "DOB_Formatted"], keep=False) & (result["Surname"] != "")
    result.loc[dup_mask, "Notes"] = result.loc[dup_mask, "Notes"].apply(lambda n: (n + "; " if n else "") + "Possible Duplicate")
    result.loc[dup_mask & (result["Flag"] == "VALID"), "Flag"] = "WARNING"

    # ── Build output Excel (exact Leadway template: 26 columns) ──
    REPORTS_DIR.mkdir(parents=True, exist_ok=True)

    output = pd.DataFrame({
        "Title": result["Title"],
        "Surname": result["Surname"],
        "Firstname": result["Firstname"],
        "Other Name": result["OtherName"],
        "Date of Birth\ndd/mm/ccyy": result["DOB_Formatted"],
        "Gender\nFemale or Male": result["Gender_Clean"],
        "Address \nNo commas": result.get("address", "").apply(lambda x: clean_val(x).replace(",", " ")) if "address" in result.columns else "",
        "City": result.get("city", "").apply(lambda x: clean_val(x).upper()) if "city" in result.columns else "",
        "Employee number": result.get("staff_id", "").apply(clean_val) if "staff_id" in result.columns else "",
        "Group Code": result.get("group_code", "").apply(clean_val) if "group_code" in result.columns else "",
        "Plan Code": result.get("plan_code", "").apply(clean_val) if "plan_code" in result.columns else "",
        "Member/DependantIndicator": result["Member_Dep"],
        "Relationship to main member": result["Relationship"],
        "Start date of member\ndd/mm/ccyy": result["Start_Formatted"],
        "Marital status": result["Marital_Clean"],
        "Contacts": result.get("contacts", "").apply(clean_val) if "contacts" in result.columns else "",
        "State": result.get("state", "").apply(lambda x: clean_val(x).upper()) if "state" in result.columns else "",
        "Captitation Provider": "",
        "Picture Path": "",
        "Place of birth": "",
        "Blood Group": "",
        "Size": "",
        "weight": "",
        "occupation": "",
        "contact 2": "",
        "email": result.get("email", "").apply(clean_val) if "email" in result.columns else "",
    })

    # Add validation columns at end
    output["Status"] = result["Flag"]
    output["Notes"] = result["Notes"]

    clean_client = client_name.replace(" ", "_")[:20]
    filename = f"Enrolment_{clean_client}_{session_id}.xlsx"
    filepath = REPORTS_DIR / filename

    buf = io.BytesIO()
    with pd.ExcelWriter(buf, engine="openpyxl") as writer:
        output.to_excel(writer, sheet_name="Enrolment", index=False)
        summary_data = {
            "Metric": ["Client", "Total Records", "Principals", "Spouses", "Dependants",
                        "Valid", "Warnings", "Red Flags", "Format Detected"],
            "Value": [client_name, len(output),
                len(output[output["Member/DependantIndicator"] == "O"]),
                len(output[output["Relationship to main member"] == "SPOUSE"]),
                len(output[output["Relationship to main member"].isin(["SON", "DAUGHTER", "DEPENDENT"])]),
                len(output[output["Status"] == "VALID"]),
                len(output[output["Status"] == "WARNING"]),
                len(output[output["Status"] == "RED"]),
                "Wide Format (expanded)" if is_wide else "Standard Format"],
        }
        pd.DataFrame(summary_data).to_excel(writer, sheet_name="Summary", index=False)

    # Apply conditional formatting
    from openpyxl import load_workbook
    from openpyxl.styles import PatternFill, Font
    buf.seek(0)
    wb = load_workbook(buf)
    ws = wb["Enrolment"]
    red_fill = PatternFill(start_color="FFCCCC", end_color="FFCCCC", fill_type="solid")
    orange_fill = PatternFill(start_color="FFE5CC", end_color="FFE5CC", fill_type="solid")
    green_fill = PatternFill(start_color="E6FFE6", end_color="E6FFE6", fill_type="solid")
    header_fill = PatternFill(start_color="262626", end_color="262626", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True, size=10)

    # Style header
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font

    # Style rows — Status column is 27 (26 template cols + Status)
    status_col = 27
    for row_idx in range(2, ws.max_row + 1):
        status = ws.cell(row=row_idx, column=status_col).value
        fill = None
        if status == "RED": fill = red_fill
        elif status == "WARNING": fill = orange_fill
        elif status == "VALID": fill = green_fill
        if fill:
            for col in range(1, ws.max_column + 1):
                ws.cell(row=row_idx, column=col).fill = fill

    # Auto-width columns
    for col in ws.columns:
        max_len = max(len(str(cell.value or "")) for cell in col)
        ws.column_dimensions[col[0].column_letter].width = min(max_len + 4, 40)

    buf2 = io.BytesIO()
    wb.save(buf2)
    buf2.seek(0)
    filepath.write_bytes(buf2.getvalue())

    # Stats for flash message
    total = len(output)
    reds = len(output[output["Status"] == "RED"])
    warns = len(output[output["Status"] == "WARNING"])
    fmt_type = "Wide format expanded" if is_wide else "Standard format"
    flash(f"Done! {total} members processed ({fmt_type}). {reds} red flags, {warns} warnings.")
    return redirect(url_for("enrolment_generator", download=filename))


# ═══════════════════════════════════════════════
# Shared — Report viewing & download
# ═══════════════════════════════════════════════

@app.route("/report/<filename>")
@login_required
def view_report(filename):
    return send_file(REPORTS_DIR / filename)


@app.route("/download/<filename>")
@login_required
def download_report(filename):
    return send_file(REPORTS_DIR / filename, as_attachment=True)


@app.route("/features-doc")
def features_doc():
    return render_template("features_doc.html")


if __name__ == "__main__":
    app.run(host="0.0.0.0", port=5000, debug=True)
